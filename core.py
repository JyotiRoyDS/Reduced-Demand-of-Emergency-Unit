import pandas as pd
import openpyxl
import os
import re
import hashlib
import requests
import numpy as np
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Any
import warnings
from datetime import datetime, date, timedelta
import re
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


warnings.filterwarnings('ignore')

# Constants for better maintainability
MONTH_MAPPING = {
    'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
    'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
    'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
}

MONTH_NUMBERS = {
    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
    'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
    'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
}

# Season mappings
SEASON_MAPPING = {
    'Jan': 'Winter', 'Feb': 'Winter', 'Mar': 'Spring',
    'Apr': 'Spring', 'May': 'Spring', 'Jun': 'Summer',
    'Jul': 'Summer', 'Aug': 'Summer', 'Sep': 'Autumn',
    'Oct': 'Autumn', 'Nov': 'Autumn', 'Dec': 'Winter'
}

# Quarter mappings
QUARTER_MAPPING = {
    'Jan': 'Q1', 'Feb': 'Q1', 'Mar': 'Q1',
    'Apr': 'Q2', 'May': 'Q2', 'Jun': 'Q2',
    'Jul': 'Q3', 'Aug': 'Q3', 'Sep': 'Q3',
    'Oct': 'Q4', 'Nov': 'Q4', 'Dec': 'Q4'
}
# ADD THESE NEW CONSTANTS FOR MONTH FILTERING
MONTH_NAME_TO_NUMBER = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

MONTH_NUMBER_TO_ABBR = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
    5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
    9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
}

# ADD THESE NEW CONSTANTS FOR DAY FILTERING
DAY_FILTER_PRESETS = {
    'first_week': {'name': 'First Week of Month', 'days': 7},
    'first_half': {'name': 'First Half of Month', 'days': 15},
    'last_week': {'name': 'Last Week of Month', 'days': 7},
    'weekend_only': {'name': 'Weekends Only', 'description': 'Saturday and Sunday'},
    'weekday_only': {'name': 'Weekdays Only', 'description': 'Monday to Friday'}
}

COMMON_DAY_RANGES = {
    'week': 7,
    'fortnight': 14,
    'month': 30,
    'quarter': 90
}

# Compiled regex patterns for better performance
PERIOD_PATTERNS = [
    re.compile(r'Period:\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*to\s*(\d{1,2})/(\d{1,2})/(\d{4})', re.IGNORECASE),
    re.compile(r'Period:\s*(\d{4})\s*-\s*(\d{4})\s+([A-Za-z]+)\s*-\s*([A-Za-z]+)', re.IGNORECASE),
    re.compile(r'Period:\s*(\d{4})\s*-\s*(\d{4})', re.IGNORECASE),
    re.compile(r'\b(20\d{2})\b')
]

DATE_PATTERN = re.compile(r'^\d{1,2}-[A-Za-z]{3}$')

SKIP_KEYWORDS = {
    'total', 'average', 'avg', 'max day', 'maximum', 'sum', 'mean', 'overall',
    'per day', 'in period', 'subtotal', 'grand total'
}

HOSPITAL_KEYWORDS = {'Hospital', 'University', 'Llandough'}


# ============================================================================
# WEATHER API INTEGRATION FUNCTIONS
# ============================================================================

def get_coordinates_for_location(location_name: str) -> Tuple[Optional[float], Optional[float], str]:
    """
    Get coordinates for a location using OpenStreetMap Nominatim API (no API key required)
    """
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': location_name,
            'format': 'json',
            'limit': 1,
            'countrycodes': 'gb'  # Limit to UK, can be modified for other countries
        }

        headers = {
            'User-Agent': 'Weather-Data-Processor/1.0 (Research Application)'
        }

        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()

        data = response.json()

        if data and len(data) > 0:
            result = data[0]
            lat = float(result['lat'])
            lon = float(result['lon'])
            address = result.get('display_name', location_name)

            print(f"🌍 Found coordinates for {location_name}: ({lat:.4f}, {lon:.4f})")
            return lat, lon, address
        else:
            print(f"⚠️ No coordinates found for {location_name}")
            return None, None, location_name

    except Exception as e:
        print(f"⚠️ Error geocoding {location_name}: {str(e)}")
        return None, None, location_name


def detect_location_from_data(df: pd.DataFrame, custom_indicators: List[str] = None) -> str:
    """
    Dynamically detect location from the data content
    """
    potential_locations = []

    # Default location indicators (can be customized)
    default_location_indicators = [
        r'cardiff',
        r'wales',
        r'uhw',
        r'university\s+hospital\s+wales',
        r'cardiff\s+and\s+vale'
    ]

    # Use custom indicators if provided, otherwise use defaults
    location_indicators = custom_indicators if custom_indicators else default_location_indicators

    # Search through the dataframe for location indicators
    search_rows = min(50, len(df))  # Check first 50 rows
    for col in df.columns:
        for row_idx in range(search_rows):
            try:
                cell_value = str(df.iloc[row_idx, col]).lower()
                for pattern in location_indicators:
                    if re.search(pattern, cell_value):
                        if 'cardiff' in cell_value or 'wales' in cell_value:
                            potential_locations.append('Cardiff, Wales, UK')
                        elif 'uhw' in cell_value:
                            potential_locations.append('Cardiff, Wales, UK')
                        # Add more location mappings as needed
            except:
                continue

    # Return the most likely location
    if potential_locations:
        return potential_locations[0]
    else:
        # Fallback to a default location
        return 'Cardiff, Wales, UK'


def fetch_weather_data(start_date: datetime, end_date: datetime, latitude: float, longitude: float,
                       timezone: str = 'Europe/London') -> pd.DataFrame:
    """Backward compatibility wrapper for robust weather fetching"""
    return fetch_weather_data_robust(start_date, end_date, latitude, longitude, timezone)
    """
    Fetch weather data from Open-Meteo API for specified date range and location
    """
    print(f"🌤️ Fetching weather data for coordinates ({latitude:.4f}, {longitude:.4f})")
    print(f"   Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    try:
        # Open-Meteo Historical Weather API
        url = "https://archive-api.open-meteo.com/v1/archive"

        params = {
            'latitude': latitude,
            'longitude': longitude,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'daily': [
                'temperature_2m_max',
                'temperature_2m_min',
                'temperature_2m_mean',
                'precipitation_sum',
                'weathercode',
                'windspeed_10m_max',
                'relative_humidity_2m_mean',
                'surface_pressure_mean'
            ],
            'timezone': timezone
        }

        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()

        data = response.json()

        if 'daily' in data:
            weather_df = pd.DataFrame({
                'Date': pd.to_datetime(data['daily']['time']),
                'Temperature_Max_C': data['daily']['temperature_2m_max'],
                'Temperature_Min_C': data['daily']['temperature_2m_min'],
                'Temperature_Mean_C': data['daily']['temperature_2m_mean'],
                'Precipitation_mm': data['daily']['precipitation_sum'],
                'Weather_Code': data['daily']['weathercode'],
                'Wind_Speed_kmh': data['daily']['windspeed_10m_max'],
                'Humidity_Percent': data['daily']['relative_humidity_2m_mean'],
                'Pressure_hPa': data['daily']['surface_pressure_mean']
            })

            # Add weather condition descriptions
            weather_df['Weather_Condition'] = weather_df['Weather_Code'].apply(decode_weather_code)

            # Add weather categories for analysis
            weather_df['Weather_Category'] = weather_df['Weather_Condition'].apply(categorize_weather)

            # Add temperature categories
            weather_df['Temperature_Category'] = weather_df['Temperature_Mean_C'].apply(categorize_temperature)

            # Add precipitation categories
            weather_df['Precipitation_Category'] = weather_df['Precipitation_mm'].apply(categorize_precipitation)

            print(f"✅ Retrieved weather data for {len(weather_df)} days")
            print(
                f"   Temperature range: {weather_df['Temperature_Mean_C'].min():.1f}°C to {weather_df['Temperature_Mean_C'].max():.1f}°C")
            print(f"   Total precipitation: {weather_df['Precipitation_mm'].sum():.1f}mm")

            # Show sample weather conditions
            sample_conditions = weather_df['Weather_Condition'].value_counts().head(3)
            print("Most common weather conditions:")
            for condition, count in sample_conditions.items():
                print(f"   - {condition}: {count} days")

            return weather_df
        else:
            print("⚠️ No weather data received from API")
            return pd.DataFrame()

    except requests.RequestException as e:
        print(f"⚠️ Failed to fetch weather data: {e}")
        return pd.DataFrame()
    except Exception as e:
        print(f"⚠️ Error processing weather data: {e}")
        return pd.DataFrame()


def decode_weather_code(code: int) -> str:
    """
    Decode WMO weather codes to human-readable descriptions
    """
    if pd.isna(code):
        return 'Unknown'

    code = int(code)

    weather_codes = {
        0: 'Clear sky',
        1: 'Mainly clear',
        2: 'Partly cloudy',
        3: 'Overcast',
        45: 'Fog',
        48: 'Depositing rime fog',
        51: 'Light drizzle',
        53: 'Moderate drizzle',
        55: 'Dense drizzle',
        56: 'Light freezing drizzle',
        57: 'Dense freezing drizzle',
        61: 'Slight rain',
        63: 'Moderate rain',
        65: 'Heavy rain',
        66: 'Light freezing rain',
        67: 'Heavy freezing rain',
        71: 'Slight snow',
        73: 'Moderate snow',
        75: 'Heavy snow',
        77: 'Snow grains',
        80: 'Slight rain showers',
        81: 'Moderate rain showers',
        82: 'Violent rain showers',
        85: 'Slight snow showers',
        86: 'Heavy snow showers',
        95: 'Thunderstorm',
        96: 'Thunderstorm with slight hail',
        99: 'Thunderstorm with heavy hail'
    }

    return weather_codes.get(code, f'Unknown ({code})')


def categorize_weather(condition: str) -> str:
    """
    Categorize weather conditions into broader categories for analysis
    """
    condition_lower = condition.lower()

    if any(word in condition_lower for word in ['clear', 'sunny']):
        return 'Clear'
    elif any(word in condition_lower for word in ['cloudy', 'overcast', 'partly']):
        return 'Cloudy'
    elif any(word in condition_lower for word in ['rain', 'drizzle', 'shower']):
        return 'Rainy'
    elif any(word in condition_lower for word in ['snow', 'sleet']):
        return 'Snowy'
    elif any(word in condition_lower for word in ['thunderstorm', 'storm']):
        return 'Stormy'
    elif any(word in condition_lower for word in ['fog', 'mist']):
        return 'Foggy'
    else:
        return 'Other'


def categorize_temperature(temp_c: float) -> str:
    """
    Categorize temperature into ranges for analysis
    """
    if pd.isna(temp_c):
        return 'Unknown'

    if temp_c < 0:
        return 'Freezing'
    elif temp_c < 5:
        return 'Very Cold'
    elif temp_c < 10:
        return 'Cold'
    elif temp_c < 15:
        return 'Cool'
    elif temp_c < 20:
        return 'Mild'
    elif temp_c < 25:
        return 'Warm'
    elif temp_c < 30:
        return 'Hot'
    else:
        return 'Very Hot'


def categorize_precipitation(precip_mm: float) -> str:
    """
    Categorize precipitation into ranges for analysis
    """
    if pd.isna(precip_mm):
        return 'Unknown'

    if precip_mm == 0:
        return 'None'
    elif precip_mm < 2.5:
        return 'Light'
    elif precip_mm < 10:
        return 'Moderate'
    elif precip_mm < 50:
        return 'Heavy'
    else:
        return 'Very Heavy'


def add_weather_info(df: pd.DataFrame, weather_df: pd.DataFrame, date_column: str = 'Date') -> pd.DataFrame:
    """
    Add weather information to the main dataframe
    """
    if weather_df.empty:
        # Add empty weather columns if no weather data
        weather_columns = [
            'Temperature_Max_C', 'Temperature_Min_C', 'Temperature_Mean_C',
            'Precipitation_mm', 'Weather_Condition', 'Weather_Category',
            'Temperature_Category', 'Precipitation_Category',
            'Wind_Speed_kmh', 'Humidity_Percent', 'Pressure_hPa'
        ]
        for col in weather_columns:
            df[col] = None
        return df

    print(f" Adding weather information to {len(df)} records")

    # Ensure both dataframes have proper date columns
    df['Date_for_Weather'] = pd.to_datetime(df[date_column], format='%d/%m/%Y', errors='coerce').dt.date
    weather_df['Date_for_Weather'] = pd.to_datetime(weather_df['Date']).dt.date

    # Merge weather data
    df_with_weather = df.merge(
        weather_df.drop('Date', axis=1),  # Drop original Date column to avoid conflicts
        on='Date_for_Weather',
        how='left'
    )

    # Clean up temporary column
    df_with_weather = df_with_weather.drop('Date_for_Weather', axis=1)

    # Count how many records got weather data
    weather_records = df_with_weather['Weather_Condition'].notna().sum()
    print(f"Added weather data to {weather_records:,} records")

    if weather_records > 0:
        # Show sample weather statistics
        print(
            f"Temperature range: {df_with_weather['Temperature_Mean_C'].min():.1f}°C to {df_with_weather['Temperature_Mean_C'].max():.1f}°C")
        print(
            f"Precipitation range: {df_with_weather['Precipitation_mm'].min():.1f}mm to {df_with_weather['Precipitation_mm'].max():.1f}mm")

        # Show weather category distribution
        weather_dist = df_with_weather['Weather_Category'].value_counts()
        print("Weather distribution:")
        for category, count in weather_dist.head(5).items():
            if pd.notna(category):
                print(f"   - {category}: {count:,} records")

    return df_with_weather


def enhance_dataframe_with_weather_data(df: pd.DataFrame, location_name: str = None,
                                        date_column: str = 'Date',
                                        custom_indicators: List[str] = None) -> pd.DataFrame:
    """Backward compatibility wrapper for graceful weather enhancement"""
    return enhance_dataframe_with_weather_graceful(df, location_name, date_column, custom_indicators)
    """
    Convenience function to enhance a DataFrame with weather data
    """
    if df.empty or date_column not in df.columns:
        print("⚠️ DataFrame is empty or missing date column")
        return df

    try:
        # Auto-detect location if not provided
        if not location_name:
            location_name = detect_location_from_data(df, custom_indicators)
            print(f"🌍 Auto-detected location: {location_name}")
        else:
            print(f"🌍 Using provided location: {location_name}")

        # Get coordinates for location
        lat, lon, address = get_coordinates_for_location(location_name)

        if lat is None or lon is None:
            print("⚠️ Could not get coordinates for location")
            return df

        print(f"📍 Location: {address}")
        print(f"📍 Coordinates: ({lat:.4f}, {lon:.4f})")

        # Get date range from DataFrame
        try:
            dates = pd.to_datetime(df[date_column], format='%d/%m/%Y', errors='coerce')
            valid_dates = dates.dropna()

            if valid_dates.empty:
                print("⚠️ No valid dates found in DataFrame")
                return df

            start_date = valid_dates.min()
            end_date = valid_dates.max()

            print(f"📅 Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        except Exception as e:
            print(f"⚠️ Error parsing dates: {e}")
            return df

        # Load weather data
        weather_df = fetch_weather_data(start_date, end_date, lat, lon)

        if weather_df.empty:
            print("⚠️ No weather data retrieved")
            return df

        # Enhance DataFrame
        return add_weather_info(df, weather_df, date_column)

    except Exception as e:
        print(f"⚠️ Error enhancing DataFrame with weather data: {e}")
        return df# ============================================================================
# NEW MONTH FILTERING FUNCTIONS
# ============================================================================

def create_month_filter(start_month: int, start_year: int, end_month: int, end_year: int) -> Dict[str, Any]:
    """
    Create a month filter configuration for precise date range filtering
    """
    from datetime import datetime, date, timedelta

    # Validate inputs
    if not (1 <= start_month <= 12) or not (1 <= end_month <= 12):
        raise ValueError("Months must be between 1 and 12")

    if start_year > end_year:
        raise ValueError("Start year cannot be greater than end year")

    if start_year == end_year and start_month > end_month:
        raise ValueError("Start month cannot be greater than end month in the same year")

    # Create start and end dates
    start_date = date(start_year, start_month, 1)

    # Calculate the last day of the end month
    if end_month == 12:
        next_month_year = end_year + 1
        next_month = 1
    else:
        next_month_year = end_year
        next_month = end_month + 1

    # Last day of end month is day before first day of next month
    end_date = date(next_month_year, next_month, 1) - timedelta(days=1)

    # Create human-readable description
    start_abbr = MONTH_NUMBER_TO_ABBR[start_month]
    end_abbr = MONTH_NUMBER_TO_ABBR[end_month]

    if start_year == end_year:
        description = f"{start_abbr} {start_year} to {end_abbr} {end_year}"
    else:
        description = f"{start_abbr} {start_year} to {end_abbr} {end_year}"

    return {
        'start_month': start_month,
        'start_year': start_year,
        'end_month': end_month,
        'end_year': end_year,
        'start_date': start_date,
        'end_date': end_date,
        'description': description,
        'type': 'month_range'
    }


# ============================================================================
# DAY FILTERING FUNCTIONS - NEW ENHANCEMENT
# ============================================================================

def create_day_filter(start_day: int, start_month: int, start_year: int,
                      end_day: int, end_month: int, end_year: int) -> Dict[str, Any]:
    """
    Create a day-level filter configuration for precise date range filtering
    """
    # Validate inputs
    if not (1 <= start_month <= 12) or not (1 <= end_month <= 12):
        raise ValueError("Months must be between 1 and 12")

    if not (1 <= start_day <= 31) or not (1 <= end_day <= 31):
        raise ValueError("Days must be between 1 and 31")

    if start_year > end_year:
        raise ValueError("Start year cannot be greater than end year")

    # Create actual date objects and validate them
    try:
        start_date = date(start_year, start_month, start_day)
        end_date = date(end_year, end_month, end_day)
    except ValueError as e:
        raise ValueError(f"Invalid date: {e}")

    if start_date > end_date:
        raise ValueError("Start date cannot be greater than end date")

    # Create human-readable description
    start_str = start_date.strftime('%d %b %Y')
    end_str = end_date.strftime('%d %b %Y')
    description = f"{start_str} to {end_str}"

    # Calculate duration
    duration_days = (end_date - start_date).days + 1

    return {
        'start_day': start_day,
        'start_month': start_month,
        'start_year': start_year,
        'end_day': end_day,
        'end_month': end_month,
        'end_year': end_year,
        'start_date': start_date,
        'end_date': end_date,
        'description': description,
        'duration_days': duration_days,
        'type': 'day_range'
    }


def date_falls_in_day_filter(check_date: date, day_filter: Dict[str, Any]) -> bool:
    """
    Check if a date falls within the day filter range
    """
    if not day_filter or day_filter.get('type') != 'day_range':
        return True  # No filter means include all

    return day_filter['start_date'] <= check_date <= day_filter['end_date']


def parse_date_for_day_filter(date_string: str, year: str) -> Optional[date]:
    """
    Parse a date string for day filtering (enhanced version)
    """
    try:
        if not year or year == 'Unknown' or not year.isdigit():
            return None

        if '-' not in date_string:
            return None

        parts = date_string.split('-')
        if len(parts) != 2:
            return None

        day_str, month_str = parts
        day = int(day_str)

        month_num = MONTH_NAME_TO_NUMBER.get(month_str.lower())
        if not month_num:
            return None

        return date(int(year), month_num, day)

    except (ValueError, TypeError):
        return None

def date_falls_in_month_filter(check_date: date, month_filter: Dict[str, Any]) -> bool:
    """
    Check if a date falls within the month filter range
    """
    if not month_filter or month_filter.get('type') != 'month_range':
        return True  # No filter means include all

    return month_filter['start_date'] <= check_date <= month_filter['end_date']


def parse_date_for_month_filter(date_string: str, year: str) -> Optional[date]:
    """
    Parse a date string in format "1-Sep" with separate year for month filtering
    """
    try:
        if not year or year == 'Unknown' or not year.isdigit():
            return None

        if '-' not in date_string:
            return None

        parts = date_string.split('-')
        if len(parts) != 2:
            return None

        day_str, month_str = parts
        day = int(day_str)

        month_num = MONTH_NAME_TO_NUMBER.get(month_str.lower())
        if not month_num:
            return None

        return date(int(year), month_num, day)

    except (ValueError, TypeError):
        return None


# MAJOR OPTIMIZATION: True Early Filtering Implementation
# This replaces the current inefficient approach with real early filtering

def should_process_section_early_filter(period_info: Dict[str, Any],
                                        year_filter: tuple = None,
                                        month_filter: Dict[str, Any] = None,
                                        day_filter: Dict[str, Any] = None) -> bool:
    """
    EARLY DECISION: Check if we should even process this section based on filters
    Returns False immediately if section doesn't overlap with filter criteria
    """
    if not any([year_filter, month_filter, day_filter]):
        return True  # No filters, process everything

    if not period_info or not period_info.get('years_span'):
        return True  # Unknown period, process cautiously

    # Get section years
    section_years = []
    for year in period_info.get('years_span', []):
        if year.isdigit():
            section_years.append(int(year))

    if not section_years:
        return True  # No valid years found, process cautiously

    section_start_year = min(section_years)
    section_end_year = max(section_years)

    # Year filter check
    if year_filter:
        filter_start, filter_end = year_filter
        # Check if there's any overlap between section years and filter years
        if section_end_year < filter_start or section_start_year > filter_end:
            return False  # No overlap, skip entirely

    # Month filter check
    if month_filter:
        filter_start_year = month_filter['start_year']
        filter_end_year = month_filter['end_year']
        # Check if section years overlap with month filter years
        if section_end_year < filter_start_year or section_start_year > filter_end_year:
            return False  # No overlap, skip entirely

    # Day filter check
    if day_filter:
        filter_start_year = day_filter['start_year']
        filter_end_year = day_filter['end_year']
        # Check if section years overlap with day filter years
        if section_end_year < filter_start_year or section_start_year > filter_end_year:
            return False  # No overlap, skip entirely

    return True  # Section has potential overlap, process it


def should_process_date_early_filter(date_header: str,
                                     year_header: str,
                                     year_filter: tuple = None,
                                     month_filter: Dict[str, Any] = None,
                                     day_filter: Dict[str, Any] = None) -> bool:
    """
    EARLY DECISION: Check if we should process this specific date column
    Returns False immediately if date doesn't match filter criteria
    """
    if not any([year_filter, month_filter, day_filter]):
        return True  # No filters, process everything

    if not year_header or year_header == 'Unknown' or not year_header.isdigit():
        return True  # Unknown year, process cautiously

    year_int = int(year_header)

    # Year filter check
    if year_filter:
        if not (year_filter[0] <= year_int <= year_filter[1]):
            return False  # Year outside range, skip

    # Month filter check
    if month_filter:
        parsed_date = parse_date_for_month_filter(date_header, year_header)
        if parsed_date:
            if not date_falls_in_month_filter(parsed_date, month_filter):
                return False  # Date outside month range, skip
        else:
            # If we can't parse the date for month filter, be conservative
            if not (month_filter['start_year'] <= year_int <= month_filter['end_year']):
                return False  # Year outside month filter years, skip

    # Day filter check
    if day_filter:
        parsed_date = parse_date_for_day_filter(date_header, year_header)
        if parsed_date:
            if not date_falls_in_day_filter(parsed_date, day_filter):
                return False  # Date outside day range, skip
        else:
            # If we can't parse the date for day filter, be conservative
            if not (day_filter['start_year'] <= year_int <= day_filter['end_year']):
                return False  # Year outside day filter years, skip

    return True  # Date passes filter checks


def get_filtered_column_indices(date_headers: List[str],
                                year_headers: List[str],
                                date_start_col: int,
                                year_filter: tuple = None,
                                month_filter: Dict[str, Any] = None,
                                day_filter: Dict[str, Any] = None) -> Tuple[List[int], List[str], List[str]]:
    """
    OPTIMIZATION: Get only the column indices we need to read based on filters
    Returns: (column_indices, filtered_date_headers, filtered_year_headers)
    """
    if not any([year_filter, month_filter, day_filter]):
        # No filters - return all columns
        column_indices = list(range(date_start_col, date_start_col + len(date_headers)))
        return column_indices, date_headers, year_headers

    # Apply early filtering to determine which columns to read
    filtered_indices = []
    filtered_dates = []
    filtered_years = []

    for i, (date_header, year_header) in enumerate(zip(date_headers, year_headers)):
        if should_process_date_early_filter(date_header, year_header, year_filter, month_filter, day_filter):
            column_index = date_start_col + i
            filtered_indices.append(column_index)
            filtered_dates.append(date_header)
            filtered_years.append(year_header)

    return filtered_indices, filtered_dates, filtered_years


def extract_data_with_true_early_filtering(excel_file_path: str,
                                           year_filter: tuple = None,
                                           month_filter: Dict[str, Any] = None,
                                           day_filter: Dict[str, Any] = None) -> List[Dict[str, Any]]:
    """
    OPTIMIZED VERSION: Extract data with true early filtering applied at every step
    This dramatically reduces processing time by skipping irrelevant data immediately
    """
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)

    print(f"📁 Extracting with early filtering: {os.path.basename(excel_file_path)}")

    # Determine active filter for logging
    if day_filter:
        filter_desc = f"Day filter: {day_filter['description']}"
    elif month_filter:
        filter_desc = f"Month filter: {month_filter['description']}"
    elif year_filter:
        filter_desc = f"Year filter: {year_filter[0]}-{year_filter[1]}"
    else:
        filter_desc = "No filter"

    print(f"  🎯 {filter_desc}")

    admissions_sheets = find_admissions_sheets(workbook)
    if not admissions_sheets:
        print(f"  ❌ No admissions sheets found")
        workbook.close()
        return []

    all_structured_data = []
    sections_processed = 0
    sections_skipped = 0

    for sheet_name in admissions_sheets:
        print(f"\n  📋 Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        table_sections = find_table_sections(sheet, sheet_name)

        for section in table_sections:
            print(f"    📊 Evaluating: {section['title']}")

            # EARLY FILTER 1: Get period info and check if section should be processed
            period_info = find_period_information_robust(sheet, section, excel_file_path)
            if not period_info:
                print(f"    ❌ Skipping - no period information")
                sections_skipped += 1
                continue

            # EARLY FILTER 2: Check if this section overlaps with our filter
            if not should_process_section_early_filter(period_info, year_filter, month_filter, day_filter):
                years_span = period_info.get('years_span', [])
                print(f"    ⏭️ SKIPPING SECTION - years {years_span} outside filter range")
                sections_skipped += 1
                continue

            print(f"    ✅ Processing section - years {period_info['years_span']} overlap with filter")
            sections_processed += 1

            # Find date headers (same as before)
            date_headers = []
            date_row = None
            date_start_col = None

            for row_num in range(section['start_row'], min(section['start_row'] + 15, sheet.max_row + 1)):
                for col_num in range(3, min(50, sheet.max_column + 1)):
                    try:
                        cell_value = sheet.cell(row=row_num, column=col_num).value
                        if cell_value and DATE_PATTERN.match(str(cell_value)):
                            date_row = row_num
                            date_start_col = col_num

                            # Collect consecutive date headers
                            for date_col in range(col_num, min(sheet.max_column + 1, col_num + 500)):
                                date_cell = sheet.cell(row=row_num, column=date_col).value
                                if date_cell and DATE_PATTERN.match(str(date_cell)):
                                    date_headers.append(str(date_cell))
                                else:
                                    if len(date_headers) > 10:
                                        break
                            break
                    except:
                        continue
                if date_headers:
                    break

            if not date_headers:
                print(f"    ❌ No date headers found")
                continue

            year_headers = assign_years_to_dates_improved(date_headers, period_info, debug=False)

            # EARLY FILTER 3: Get only the column indices we need to read
            filtered_column_indices, filtered_date_headers, filtered_year_headers = get_filtered_column_indices(
                date_headers, year_headers, date_start_col, year_filter, month_filter, day_filter
            )

            original_columns = len(date_headers)
            filtered_columns = len(filtered_column_indices)

            if filtered_columns == 0:
                print(f"    ⏭️ SKIPPING SECTION - no dates match filter criteria")
                continue

            print(f"    📅 Columns: {original_columns} total → {filtered_columns} filtered ({filter_desc})")

            # Find data start row (same as before)
            data_start_row = date_row + 1 if date_row else section['start_row'] + 3

            # OPTIMIZATION: Only read the columns we actually need
            max_col_needed = max(filtered_column_indices) + 1 if filtered_column_indices else 10
            min_col_needed = min(filtered_column_indices) if filtered_column_indices else 1

            # Extract data rows - ONLY READING NEEDED COLUMNS
            current_hospital = None
            current_method_category = None
            end_row = min(data_start_row + 500, sheet.max_row + 1)
            records_extracted = 0

            for row_num in range(data_start_row, end_row):
                try:
                    # OPTIMIZATION: Only read columns from 1 to max_col_needed
                    row_values = [sheet.cell(row=row_num, column=col).value
                                  for col in range(1, max_col_needed)]

                    if not any(row_values[:4]):
                        continue

                    if should_skip_row(row_values):
                        continue

                    # Hospital and category extraction (same logic as before)
                    hospital_name = find_hospital_name(row_values)
                    if hospital_name:
                        current_hospital = hospital_name

                    if not current_hospital:
                        continue

                    # Category extraction logic (same as original)
                    if section['type'] == 'method_patient_class':
                        if hospital_name and len(row_values) > 1 and row_values[1]:
                            current_method_category = str(row_values[1]).strip()
                        method_or_category = current_method_category
                        patient_class = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else None
                        category_type = "Method"
                    elif section['type'] == 'directorate':
                        method_or_category = str(row_values[2]).strip() if len(row_values) > 2 and row_values[
                            2] else None
                        patient_class = "Emergency Inpatient"
                        category_type = "Directorate"
                    elif section['type'] == 'diagnosis':
                        method_or_category = str(row_values[2]).strip() if len(row_values) > 2 and row_values[
                            2] else None
                        patient_class = "Emergency Inpatient"
                        category_type = "Primary Diagnosis"
                    else:
                        method_or_category = str(row_values[1]).strip() if len(row_values) > 1 and row_values[
                            1] else None
                        patient_class = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else None
                        category_type = "Unknown"

                    if not method_or_category or not patient_class or not method_or_category.strip():
                        continue

                    # Skip rows with unwanted keywords
                    method_text = str(method_or_category).strip().lower()
                    patient_text = str(patient_class).strip().lower()
                    if any(indicator in method_text or indicator in patient_text for indicator in SKIP_KEYWORDS):
                        continue

                    # Create record
                    record = {
                        'hospital': current_hospital,
                        'section': section['title'],
                        'category_type': category_type,
                        'category_value': method_or_category,
                        'patient_class': patient_class,
                        'sheet_name': sheet_name,
                        'file_name': os.path.basename(excel_file_path),
                        'period_info': period_info,
                        'daily_data': {}
                    }

                    # OPTIMIZATION: Only extract values from filtered columns
                    data_points_added = 0
                    for i, col_index in enumerate(filtered_column_indices):
                        if col_index - 1 < len(row_values):  # col_index is 1-based, row_values is 0-based
                            value = row_values[col_index - 1]
                            if value is not None and str(value).strip():
                                try:
                                    if isinstance(value, str):
                                        value = value.replace(',', '').replace(' ', '')
                                    num_value = int(float(value))

                                    date_header = filtered_date_headers[i]
                                    year_header = filtered_year_headers[i]
                                    date_key = f"{date_header}-{year_header}" if year_header and year_header != 'Unknown' else date_header

                                    record['daily_data'][date_key] = num_value
                                    data_points_added += 1
                                except (ValueError, TypeError):
                                    pass

                    if record['daily_data']:
                        all_structured_data.append(record)
                        records_extracted += 1

                except Exception as e:
                    continue

            print(f"    ✅ Extracted {records_extracted} records from this section")

    workbook.close()

    total_data_points = sum(len(record['daily_data']) for record in all_structured_data)
    print(f"  ✅ OPTIMIZATION RESULTS:")
    print(f"    📊 Sections processed: {sections_processed}")
    print(f"    ⏭️ Sections skipped: {sections_skipped}")
    print(f"    📋 Total records: {len(all_structured_data)}")
    print(f"    📊 Total data points: {total_data_points}")
    print(f"    ⚡ Filter applied: {filter_desc}")

    return all_structured_data


# ============================================================================
# EXISTING CORE FUNCTIONS (keeping all the original functionality)
# ============================================================================

def fetch_bank_holidays(years: List[int], region: str = "england-and-wales", cache_timeout: int = 3600) -> pd.DataFrame:
    """
    Fetch bank holidays from GOV.UK API for specified years.
    """
    print(f"🏛️ Fetching bank holidays for years: {years}")

    try:
        url = "https://www.gov.uk/bank-holidays.json"
        response = requests.get(url, timeout=10)
        response.raise_for_status()

        data = response.json()
        events = data[region]["events"]

        # Convert to DataFrame
        df = pd.DataFrame(events)
        if df.empty:
            print(f"⚠️ No bank holiday data received from API")
            return pd.DataFrame(columns=["Date", "Event", "Type", "Audience"])

        # Process the data
        df["Date"] = pd.to_datetime(df["date"], format="%Y-%m-%d")
        df["Event"] = df["title"]
        df["Type"] = "Bank Holiday"
        df["Audience"] = "General Public"

        # Filter to requested years
        df = df[df["Date"].dt.year.isin(years)]
        df = df[["Date", "Event", "Type", "Audience"]].copy()

        print(f"✅ Retrieved {len(df)} bank holidays for {region}")
        return df

    except requests.RequestException as e:
        print(f"⚠️ Failed to fetch bank holidays from GOV.UK API: {e}")
        print(f"   Continuing without bank holiday data...")
        return pd.DataFrame(columns=["Date", "Event", "Type", "Audience"])
    except Exception as e:
        print(f"⚠️ Error processing bank holiday data: {e}")
        return pd.DataFrame(columns=["Date", "Event", "Type", "Audience"])


def parse_period_text_improved(period_text: str) -> Optional[Dict[str, Any]]:
    """Parse period text based on actual patterns found in the Excel files"""
    if not period_text:
        return None

    print(f"    🔍 Parsing period text: '{period_text}'")

    # Pattern 1: "Period: 01/09/2021 to 31/03/2022"
    match1 = PERIOD_PATTERNS[0].search(period_text)
    if match1:
        start_day, start_month, start_year = match1.groups()[:3]
        end_day, end_month, end_year = match1.groups()[3:]

        print(f"    ✅ Pattern 1 matched: {start_day}/{start_month}/{start_year} to {end_day}/{end_month}/{end_year}")

        return {
            'start_date': f"{start_day.zfill(2)}/{start_month.zfill(2)}/{start_year}",
            'end_date': f"{end_day.zfill(2)}/{end_month.zfill(2)}/{end_year}",
            'start_year': start_year,
            'end_year': end_year,
            'years_span': [start_year, end_year] if start_year != end_year else [start_year],
            'cross_year': start_year != end_year,
            'period_type': 'explicit_dates',
            'confidence': 'high',
            'source_text': period_text.strip()
        }

    # Pattern 2: "Period: 2021 - 2024   March - September"
    match2 = PERIOD_PATTERNS[1].search(period_text)
    if match2:
        start_year, end_year, start_month, end_month = match2.groups()
        years_span = [str(year) for year in range(int(start_year), int(end_year) + 1)]

        print(f"    ✅ Pattern 2 matched: {start_year}-{end_year}, {start_month}-{end_month}")

        return {
            'start_date': None,
            'end_date': None,
            'start_year': start_year,
            'end_year': end_year,
            'start_month': start_month,
            'end_month': end_month,
            'years_span': years_span,
            'cross_year': True,
            'multi_year_span': True,
            'period_type': 'year_range_with_months',
            'confidence': 'high',
            'source_text': period_text.strip()
        }

    # Pattern 3: Simple year range "Period: 2022 - 2023"
    match3 = PERIOD_PATTERNS[2].search(period_text)
    if match3:
        start_year, end_year = match3.groups()
        years_span = [str(year) for year in range(int(start_year), int(end_year) + 1)]

        print(f"    ✅ Pattern 3 matched: {start_year}-{end_year}")

        return {
            'start_date': None,
            'end_date': None,
            'start_year': start_year,
            'end_year': end_year,
            'years_span': years_span,
            'cross_year': True,
            'period_type': 'year_range',
            'confidence': 'medium',
            'source_text': period_text.strip()
        }

    # Fallback: extract all years found
    years = PERIOD_PATTERNS[3].findall(period_text)
    if years:
        years = sorted(list(set(years)))
        print(f"    ⚠️ Fallback pattern: extracted years {years}")

        return {
            'start_date': None,
            'end_date': None,
            'start_year': years[0],
            'end_year': years[-1],
            'years_span': years,
            'cross_year': len(years) > 1,
            'period_type': 'extracted_years',
            'confidence': 'low',
            'source_text': period_text.strip()
        }

    print(f"    ❌ No patterns matched for: '{period_text}'")
    return None


def find_period_information_robust(sheet, section: Dict[str, Any], excel_file_path: str) -> Optional[Dict[str, Any]]:
    """Robust period detection based on actual Excel file structures"""
    print(f"  🔍 Searching for period information...")

    search_locations = [
        {'rows': range(1, 26), 'cols': range(1, 3), 'priority': 'high'},
        {'rows': range(26, 51), 'cols': range(1, 6), 'priority': 'medium'},
        {'rows': range(max(1, section['start_row'] - 10), min(section['start_row'] + 10, 101)),
         'cols': range(1, 6), 'priority': 'medium'}
    ]

    for search_area in search_locations:
        for row_num in search_area['rows']:
            for col_num in search_area['cols']:
                try:
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    if cell_value and isinstance(cell_value, str):
                        cell_text = cell_value.strip()

                        if cell_text.lower().startswith('period:'):
                            print(f"  ✅ Found at Row {row_num}, Col {col_num}: '{cell_text}'")

                            period_info = parse_period_text_improved(cell_text)
                            if period_info:
                                period_info['found_location'] = f"Row {row_num}, Col {col_num}"
                                period_info['search_priority'] = search_area['priority']
                                return period_info

                except Exception:
                    continue

    print(f"  ❌ No period information found - this is unusual!")
    return None


def assign_years_to_dates_improved(date_headers: List[str], period_info: Dict[str, Any], debug: bool = True) -> List[
    Optional[str]]:
    """Improved year assignment based on detected period information"""
    if not date_headers or not period_info:
        return [None] * len(date_headers)

    if debug:
        print(f"  📅 Assigning years to {len(date_headers)} dates")
        print(f"    Period type: {period_info.get('period_type', 'unknown')}")
        print(f"    Years available: {period_info.get('years_span', [])}")

    years_span = period_info.get('years_span', [])
    period_type = period_info.get('period_type', 'unknown')

    if not years_span:
        if debug:
            print(f"    ❌ No years available in period info")
        return [None] * len(date_headers)

    # Create a more efficient year assignment function
    def get_month_from_date(date_str: str) -> str:
        return date_str.split('-')[1] if '-' in date_str else ''

    def assign_year_by_month(month: str, start_year: str, end_year: str) -> str:
        if month in ['Sep', 'Oct', 'Nov', 'Dec']:
            return start_year
        elif month in ['Jan', 'Feb', 'Mar']:
            return end_year
        else:
            return start_year  # Default fallback

    year_assignments = []

    if period_type == 'explicit_dates':
        start_year = period_info['start_year']
        end_year = period_info['end_year']

        if start_year == end_year:
            year_assignments = [start_year] * len(date_headers)
        else:
            for date in date_headers:
                month = get_month_from_date(date)
                year_assignments.append(assign_year_by_month(month, start_year, end_year))

    elif period_type == 'year_range_with_months':
        if len(years_span) == 1:
            year_assignments = [years_span[0]] * len(date_headers)
        else:
            dates_per_year = len(date_headers) / len(years_span)
            for i in range(len(date_headers)):
                year_index = min(int(i / dates_per_year), len(years_span) - 1)
                year_assignments.append(years_span[year_index])

    elif period_type == 'year_range':
        if len(years_span) == 1:
            year_assignments = [years_span[0]] * len(date_headers)
        elif len(years_span) == 2:
            first_year, second_year = years_span[0], years_span[1]
            for date in date_headers:
                month = get_month_from_date(date)
                year_assignments.append(assign_year_by_month(month, first_year, second_year))
        else:
            dates_per_year = len(date_headers) / len(years_span)
            for i in range(len(date_headers)):
                year_index = min(int(i / dates_per_year), len(years_span) - 1)
                year_assignments.append(years_span[year_index])
    else:
        # Fallback
        if len(years_span) == 1:
            year_assignments = [years_span[0]] * len(date_headers)
        else:
            dates_per_year = len(date_headers) / len(years_span)
            for i in range(len(date_headers)):
                year_index = min(int(i / dates_per_year), len(years_span) - 1)
                year_assignments.append(years_span[year_index])

    if debug:
        year_counts = {}
        for year in year_assignments:
            year_counts[year] = year_counts.get(year, 0) + 1
        print(f"    ✅ Year distribution: {year_counts}")

    return year_assignments


def find_admissions_sheets(workbook) -> List[str]:
    """Find all sheets that contain admissions data"""
    return [sheet_name for sheet_name in workbook.sheetnames
            if 'Admissions' in sheet_name or 'admissions' in sheet_name]


def determine_table_type(title: str) -> str:
    """Determine the table type based on the title"""
    title_lower = title.lower()
    if 'method' in title_lower and 'patient class' in title_lower:
        return 'method_patient_class'
    elif 'directorate' in title_lower:
        return 'directorate'
    elif 'diagnosis' in title_lower:
        return 'diagnosis'
    else:
        return 'unknown'


def find_table_sections(sheet, sheet_name: str) -> List[Dict[str, Any]]:
    """Find all table sections in a sheet"""
    table_sections = []

    for row_num in range(1, min(101, sheet.max_row + 1)):
        for col_num in range(1, 6):
            try:
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if (cell_value and isinstance(cell_value, str) and
                        'Admissions' in cell_value and 'by' in cell_value):
                    table_sections.append({
                        'title': cell_value,
                        'start_row': row_num,
                        'start_col': col_num,
                        'type': determine_table_type(cell_value),
                        'sheet_name': sheet_name
                    })
                    break
            except:
                continue

    return table_sections


def create_date_key_signature(date_key: str, hospital: str, category_type: str,
                              category_value: str, patient_class: str) -> str:
    """Create a unique signature for a data point to detect cross-file overlaps"""
    key_components = [
        str(date_key).strip().lower(),
        str(hospital).strip().lower(),
        str(category_type).strip().lower(),
        str(category_value).strip().lower(),
        str(patient_class).strip().lower()
    ]

    signature_string = '|'.join(key_components)
    return hashlib.md5(signature_string.encode()).hexdigest()


def should_skip_row(row_values: List[Any]) -> bool:
    """Check if a row should be skipped based on skip keywords"""
    for i in range(min(5, len(row_values))):
        if row_values[i]:
            cell_text = str(row_values[i]).strip().lower()
            if any(keyword in cell_text for keyword in SKIP_KEYWORDS):
                return True
    return False


def find_hospital_name(row_values: List[Any]) -> Optional[str]:
    """Extract hospital name from row values"""
    for i in range(min(2, len(row_values))):
        if row_values[i] and str(row_values[i]).strip():
            hospital_name = str(row_values[i]).strip()
            if any(keyword in hospital_name for keyword in HOSPITAL_KEYWORDS):
                if not any(keyword in hospital_name.lower() for keyword in SKIP_KEYWORDS):
                    return hospital_name
    return None


def get_day_of_week_info(day: int, month: str, year: str) -> Dict[str, Any]:
    """Get day of week information including weekend status"""
    try:
        if year == 'Unknown' or not year.isdigit():
            return {
                'day_of_week': 'Unknown',
                'day_of_week_num': 0,
                'is_weekend': 'Unknown'
            }

        month_num = int(MONTH_NUMBERS.get(month, '00'))
        if month_num == 0:
            return {
                'day_of_week': 'Unknown',
                'day_of_week_num': 0,
                'is_weekend': 'Unknown'
            }

        date_obj = date(int(year), month_num, day)
        day_of_week_num = date_obj.weekday()  # 0=Monday, 6=Sunday

        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_of_week = day_names[day_of_week_num]

        # Weekend is Saturday (5) and Sunday (6)
        is_weekend = 'Yes' if day_of_week_num in [5, 6] else 'No'

        return {
            'day_of_week': day_of_week,
            'day_of_week_num': day_of_week_num + 1,  # Convert to 1-7 (1=Monday, 7=Sunday)
            'is_weekend': is_weekend
        }
    except (ValueError, TypeError):
        return {
            'day_of_week': 'Unknown',
            'day_of_week_num': 0,
            'is_weekend': 'Unknown'
        }


def add_bank_holiday_info(df: pd.DataFrame, bank_holidays_df: pd.DataFrame) -> pd.DataFrame:
    """
    Add bank holiday information to the main dataframe.
    """
    if bank_holidays_df.empty:
        df['Is_Bank_Holiday'] = 'No'
        df['Bank_Holiday_Name'] = None
        return df

    print(f"🏛️ Adding bank holiday information to {len(df)} records")

    # Ensure Date column is datetime
    df['Date_Parsed'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
    bank_holidays_df['Date'] = pd.to_datetime(bank_holidays_df['Date'])

    # Create a lookup dictionary for faster matching
    bank_holiday_lookup = {}
    for _, row in bank_holidays_df.iterrows():
        date_key = row['Date'].strftime('%Y-%m-%d')
        bank_holiday_lookup[date_key] = row['Event']

    # Add bank holiday information
    def check_bank_holiday(date_val):
        if pd.isna(date_val):
            return 'No', None

        date_key = date_val.strftime('%Y-%m-%d')
        if date_key in bank_holiday_lookup:
            return 'Yes', bank_holiday_lookup[date_key]
        return 'No', None

    df[['Is_Bank_Holiday', 'Bank_Holiday_Name']] = df['Date_Parsed'].apply(
        lambda x: pd.Series(check_bank_holiday(x))
    )

    # Clean up temporary column
    df = df.drop('Date_Parsed', axis=1)

    bank_holiday_count = (df['Is_Bank_Holiday'] == 'Yes').sum()
    print(f"✅ Identified {bank_holiday_count} bank holiday records")

    return df


# ADD ALL THE NEW ROBUST WEATHER FUNCTIONS HERE

def fetch_weather_data_robust(start_date: datetime, end_date: datetime, latitude: float, longitude: float,
                              timezone: str = 'Europe/London') -> pd.DataFrame:
    """Robust weather data fetching with SSL error handling and multiple fallback strategies"""
    print(f"🌤️ Fetching weather data for coordinates ({latitude:.4f}, {longitude:.4f})")
    print(f"   Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    try:
        return _fetch_with_robust_session(start_date, end_date, latitude, longitude, timezone)
    except Exception as e:
        print(f"⚠️ Strategy 1 failed: {str(e)[:100]}...")

    try:
        return _fetch_with_ssl_disabled(start_date, end_date, latitude, longitude, timezone)
    except Exception as e:
        print(f"⚠️ Strategy 2 failed: {str(e)[:100]}...")

    try:
        return _fetch_with_simple_params(start_date, end_date, latitude, longitude, timezone)
    except Exception as e:
        print(f"⚠️ Strategy 3 failed: {str(e)[:100]}...")

    print("⚠️ All weather data fetching strategies failed")
    return pd.DataFrame()


def _fetch_with_robust_session(start_date: datetime, end_date: datetime, latitude: float, longitude: float,
                               timezone: str) -> pd.DataFrame:
    session = requests.Session()

    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"]
    )

    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    session.headers.update({
        'User-Agent': 'Healthcare-Data-Processor/1.0 (Research Application)',
        'Accept': 'application/json',
    })

    url = "https://archive-api.open-meteo.com/v1/archive"

    params = {
        'latitude': latitude,
        'longitude': longitude,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'daily': [
            'temperature_2m_max',
            'temperature_2m_min',
            'temperature_2m_mean',
            'precipitation_sum',
            'weathercode',
            'windspeed_10m_max',
            'relative_humidity_2m_mean',
            'surface_pressure_mean'
        ],
        'timezone': timezone
    }

    response = session.get(url, params=params, timeout=(10, 30))
    response.raise_for_status()

    return _process_weather_response(response.json())


def _fetch_with_ssl_disabled(start_date: datetime, end_date: datetime, latitude: float, longitude: float,
                             timezone: str) -> pd.DataFrame:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    url = "https://archive-api.open-meteo.com/v1/archive"

    params = {
        'latitude': latitude,
        'longitude': longitude,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'daily': [
            'temperature_2m_max',
            'temperature_2m_min',
            'temperature_2m_mean',
            'precipitation_sum',
            'weathercode',
            'windspeed_10m_max',
            'relative_humidity_2m_mean',
            'surface_pressure_mean'
        ],
        'timezone': timezone
    }

    headers = {
        'User-Agent': 'Healthcare-Data-Processor/1.0 (Research Application)'
    }

    response = requests.get(
        url,
        params=params,
        headers=headers,
        verify=False,
        timeout=(15, 45)
    )
    response.raise_for_status()

    return _process_weather_response(response.json())


def _fetch_with_simple_params(start_date: datetime, end_date: datetime, latitude: float, longitude: float,
                              timezone: str) -> pd.DataFrame:
    url = "https://archive-api.open-meteo.com/v1/archive"

    params = {
        'latitude': latitude,
        'longitude': longitude,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'daily': 'temperature_2m_mean,precipitation_sum,weathercode',
        'timezone': timezone
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (compatible; HealthcareProcessor/1.0)',
        'Accept': 'application/json'
    }

    response = requests.get(
        url,
        params=params,
        headers=headers,
        timeout=(20, 60),
        verify=False
    )

    response.raise_for_status()
    return _process_weather_response(response.json())


def _process_weather_response(data: dict) -> pd.DataFrame:
    if 'daily' not in data:
        return pd.DataFrame()

    daily_data = data['daily']

    weather_df = pd.DataFrame({
        'Date': pd.to_datetime(daily_data['time']),
    })

    weather_df['Temperature_Max_C'] = daily_data.get('temperature_2m_max', [None] * len(weather_df))
    weather_df['Temperature_Min_C'] = daily_data.get('temperature_2m_min', [None] * len(weather_df))
    weather_df['Temperature_Mean_C'] = daily_data.get('temperature_2m_mean', [None] * len(weather_df))
    weather_df['Precipitation_mm'] = daily_data.get('precipitation_sum', [None] * len(weather_df))
    weather_df['Weather_Code'] = daily_data.get('weathercode', [None] * len(weather_df))
    weather_df['Wind_Speed_kmh'] = daily_data.get('windspeed_10m_max', [None] * len(weather_df))
    weather_df['Humidity_Percent'] = daily_data.get('relative_humidity_2m_mean', [None] * len(weather_df))
    weather_df['Pressure_hPa'] = daily_data.get('surface_pressure_mean', [None] * len(weather_df))

    weather_df['Weather_Condition'] = weather_df['Weather_Code'].apply(decode_weather_code)
    weather_df['Weather_Category'] = weather_df['Weather_Condition'].apply(categorize_weather)
    weather_df['Temperature_Category'] = weather_df['Temperature_Mean_C'].apply(categorize_temperature)
    weather_df['Precipitation_Category'] = weather_df['Precipitation_mm'].apply(categorize_precipitation)

    return weather_df


def enhance_dataframe_with_weather_graceful(df: pd.DataFrame, location_name: str = None,
                                            date_column: str = 'Date',
                                            custom_indicators: List[str] = None) -> pd.DataFrame:
    if df.empty or date_column not in df.columns:
        return _add_empty_weather_columns(df)

    try:
        if not location_name:
            location_name = detect_location_from_data(df, custom_indicators)

        lat, lon, address = get_coordinates_for_location(location_name)

        if lat is None or lon is None:
            return _add_empty_weather_columns(df)

        try:
            dates = pd.to_datetime(df[date_column], format='%d/%m/%Y', errors='coerce')
            valid_dates = dates.dropna()

            if valid_dates.empty:
                return _add_empty_weather_columns(df)

            start_date = valid_dates.min()
            end_date = valid_dates.max()

        except Exception as e:
            return _add_empty_weather_columns(df)

        weather_df = fetch_weather_data_robust(start_date, end_date, lat, lon)

        if weather_df.empty:
            return _add_empty_weather_columns(df)

        return add_weather_info(df, weather_df, date_column)

    except Exception as e:
        print(f"⚠️ Weather enhancement failed: {e}")
        return _add_empty_weather_columns(df)


def _add_empty_weather_columns(df: pd.DataFrame) -> pd.DataFrame:
    weather_columns = [
        'Temperature_Max_C', 'Temperature_Min_C', 'Temperature_Mean_C',
        'Precipitation_mm', 'Weather_Condition', 'Weather_Category',
        'Temperature_Category', 'Precipitation_Category',
        'Wind_Speed_kmh', 'Humidity_Percent', 'Pressure_hPa'
    ]

    for col in weather_columns:
        if col not in df.columns:
            df[col] = None

    return df

# ADD ALL THESE SORTING FUNCTIONS AFTER add_bank_holiday_info function

def sort_dataframe_by_filter_criteria(df: pd.DataFrame,
                                      year_filter: tuple = None,
                                      month_filter: Dict[str, Any] = None,
                                      day_filter: Dict[str, Any] = None) -> pd.DataFrame:
    """Sort DataFrame according to the user's filter criteria for clear, expected ordering"""
    if df.empty:
        return df

    print(f"🔄 Sorting {len(df)} records according to filter criteria...")

    if day_filter:
        print(f"📅 Sorting by day filter: {day_filter['description']}")
        df = sort_by_day_filter(df, day_filter)
    elif month_filter:
        print(f"📅 Sorting by month filter: {month_filter['description']}")
        df = sort_by_month_filter(df, month_filter)
    elif year_filter:
        print(f"📅 Sorting by year filter: {year_filter[0]}-{year_filter[1]}")
        df = sort_by_year_filter(df, year_filter)
    else:
        print("📅 Sorting by chronological order (no filter)")
        df = sort_chronologically(df)

    print(f"✅ Sorting complete - data now in proper chronological order")
    return df


def sort_by_year_filter(df: pd.DataFrame, year_filter: tuple) -> pd.DataFrame:
    """Sort DataFrame according to year filter criteria"""
    if df.empty:
        return df

    df = df.copy()

    def create_year_sort_key(row):
        try:
            year = row.get('Year', 'Unknown')
            if year == 'Unknown' or not str(year).isdigit():
                year_num = 9999
            else:
                year_num = int(year)

            month_num_str = row.get('Month_num', '00')
            try:
                month_num = int(month_num_str) if month_num_str != '00' else 99
            except:
                month_num = 99

            date_str = row.get('Date', '')
            try:
                if '/' in date_str:
                    day = int(date_str.split('/')[0])
                else:
                    day = 1
            except:
                day = 99

            category = str(row.get('Category_Value', 'ZZZ'))
            return (year_num, month_num, day, category)
        except:
            return (9999, 99, 99, 'ZZZ')

    df['_sort_key'] = df.apply(create_year_sort_key, axis=1)
    df = df.sort_values('_sort_key').drop('_sort_key', axis=1)

    if 'Year' in df.columns:
        year_counts = df['Year'].value_counts().sort_index()
        print(f"   📊 Year distribution after sorting: {dict(year_counts)}")

    return df


def sort_by_month_filter(df: pd.DataFrame, month_filter: Dict[str, Any]) -> pd.DataFrame:
    """Sort DataFrame according to month filter criteria"""
    if df.empty:
        return df

    df = df.copy()

    def create_month_sort_key(row):
        try:
            year = row.get('Year', 'Unknown')
            month_num_str = row.get('Month_num', '00')
            date_str = row.get('Date', '')

            if year == 'Unknown' or not str(year).isdigit():
                year_num = 9999
            else:
                year_num = int(year)

            try:
                month_num = int(month_num_str) if month_num_str != '00' else 99
            except:
                month_num = 99

            try:
                if '/' in date_str:
                    day = int(date_str.split('/')[0])
                else:
                    day = 1
            except:
                day = 99

            filter_start_year = month_filter['start_year']
            filter_start_month = month_filter['start_month']

            if year_num != 9999 and month_num != 99:
                months_from_start = (year_num - filter_start_year) * 12 + (month_num - filter_start_month)
            else:
                months_from_start = 9999

            category = str(row.get('Category_Value', 'ZZZ'))
            return (months_from_start, year_num, month_num, day, category)
        except:
            return (9999, 9999, 99, 99, 'ZZZ')

    df['_sort_key'] = df.apply(create_month_sort_key, axis=1)
    df = df.sort_values('_sort_key').drop('_sort_key', axis=1)

    if 'Month' in df.columns and 'Year' in df.columns:
        month_year_dist = df.groupby(['Year', 'Month']).size().head(10)
        print(f"   📊 First 10 Month-Year combinations after sorting:")
        for (year, month), count in month_year_dist.items():
            print(f"      {month} {year}: {count} records")

    return df


def sort_by_day_filter(df: pd.DataFrame, day_filter: Dict[str, Any]) -> pd.DataFrame:
    """Sort DataFrame according to day filter criteria"""
    if df.empty:
        return df

    df = df.copy()

    def create_day_sort_key(row):
        try:
            date_str = row.get('Date', '')
            year = row.get('Year', 'Unknown')

            sort_date = None

            if '/' in date_str and len(date_str.split('/')) == 3:
                try:
                    sort_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                except:
                    pass

            if sort_date is None and year != 'Unknown':
                try:
                    month_num_str = row.get('Month_num', '00')
                    if month_num_str != '00':
                        if '/' in date_str:
                            day = int(date_str.split('/')[0])
                        else:
                            day = 1
                        month = int(month_num_str)
                        year_num = int(year)
                        sort_date = date(year_num, month, day)
                except:
                    pass

            if sort_date:
                filter_start = day_filter['start_date']
                days_from_start = (sort_date - filter_start).days

                if days_from_start < 0:
                    days_from_start = 999999

                category = str(row.get('Category_Value', 'ZZZ'))
                return (days_from_start, sort_date.year, sort_date.month, sort_date.day, category)
            else:
                return (999999, 9999, 99, 99, 'ZZZ')

        except:
            return (999999, 9999, 99, 99, 'ZZZ')

    df['_sort_key'] = df.apply(create_day_sort_key, axis=1)
    df = df.sort_values('_sort_key').drop('_sort_key', axis=1)

    if 'Date' in df.columns:
        try:
            valid_dates = df['Date'][df['Date'].str.contains(r'\d{2}/\d{2}/\d{4}', na=False)]
            if not valid_dates.empty:
                first_date = valid_dates.iloc[0]
                last_date = valid_dates.iloc[-1]
                print(f"   📊 Date range after sorting: {first_date} to {last_date}")
                print(f"   📊 Total records with valid dates: {len(valid_dates)}")
        except:
            pass

    return df


def sort_chronologically(df: pd.DataFrame) -> pd.DataFrame:
    """Sort DataFrame in pure chronological order when no filter is applied"""
    if df.empty:
        return df

    df = df.copy()

    def create_chrono_sort_key(row):
        try:
            year = row.get('Year', 'Unknown')
            month_num_str = row.get('Month_num', '00')
            date_str = row.get('Date', '')

            if year == 'Unknown' or not str(year).isdigit():
                year_num = 9999
            else:
                year_num = int(year)

            try:
                month_num = int(month_num_str) if month_num_str != '00' else 99
            except:
                month_num = 99

            try:
                if '/' in date_str:
                    day = int(date_str.split('/')[0])
                else:
                    day = 1
            except:
                day = 99

            hospital = str(row.get('Hospital_Name', 'ZZZ'))
            category = str(row.get('Category_Value', 'ZZZ'))

            return (year_num, month_num, day, hospital, category)
        except:
            return (9999, 99, 99, 'ZZZ', 'ZZZ')

    df['_sort_key'] = df.apply(create_chrono_sort_key, axis=1)
    df = df.sort_values('_sort_key').drop('_sort_key', axis=1)

    return df

def extract_basic_admissions_data_with_enhanced_filter(excel_file_path: str,
                                                       year_filter: tuple = None,
                                                       month_filter: Dict[str, Any] = None,
                                                       day_filter: Dict[str, Any] = None) -> List[Dict[str, Any]]:
    """
    REPLACEMENT FUNCTION: Now uses true early filtering for maximum efficiency
    """
    return extract_data_with_true_early_filtering(excel_file_path, year_filter, month_filter, day_filter)
    """
    ENHANCED VERSION: Extract data with year, month, OR day filtering applied during extraction
    Day filter takes precedence over month filter, which takes precedence over year filter
    """
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)

    print(f"📁 Extracting: {os.path.basename(excel_file_path)}")

    # Determine which filter to use (priority: day > month > year)
    active_filter = None
    filter_description = "No filter"

    if day_filter:
        active_filter = day_filter
        filter_description = f"Day filter: {day_filter['description']}"
        print(f"📅 {filter_description}")
    elif month_filter:
        active_filter = month_filter
        filter_description = f"Month filter: {month_filter['description']}"
        print(f"📅 {filter_description}")
    elif year_filter:
        active_filter = {'type': 'year_range', 'start_year': year_filter[0], 'end_year': year_filter[1]}
        filter_description = f"Year filter: {year_filter[0]}-{year_filter[1]}"
        print(f"📅 {filter_description}")
    else:
        print("📅 No filter applied - processing all data")

    admissions_sheets = find_admissions_sheets(workbook)
    if not admissions_sheets:
        print(f"  ❌ No admissions sheets found")
        workbook.close()
        return []

    print(f"  📊 Found {len(admissions_sheets)} admissions sheets: {admissions_sheets}")

    all_structured_data = []
    file_context = {'file_name': os.path.basename(excel_file_path)}

    for sheet_name in admissions_sheets:
        print(f"\n  📋 Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        table_sections = find_table_sections(sheet, sheet_name)

        if not table_sections:
            print(f"    ❌ No table sections found in {sheet_name}")
            continue

        for section in table_sections:
            print(f"    📊 Processing: {section['type']} - {section['title']}")

            period_info = find_period_information_robust(sheet, section, excel_file_path)

            if not period_info:
                print(f"    ❌ Skipping section - no period information found")
                continue

            # EARLY FILTER: Check if this section overlaps with our filter
            if active_filter and active_filter['type'] == 'year_range':
                years_span = period_info.get('years_span', [])
                section_years = [int(year) for year in years_span if year.isdigit()]

                if not any(active_filter['start_year'] <= year <= active_filter['end_year'] for year in section_years):
                    print(f"    ⏭️ Skipping section - years {section_years} outside filter range")
                    continue

            elif active_filter and active_filter['type'] == 'month_range':
                years_span = period_info.get('years_span', [])
                section_years = [int(year) for year in years_span if year.isdigit()]

                # Check if section years overlap with month filter years
                filter_years = list(range(active_filter['start_year'], active_filter['end_year'] + 1))
                if not any(year in filter_years for year in section_years):
                    print(f"    ⏭️ Skipping section - years {section_years} outside month filter years {filter_years}")
                    continue

            print(f"    ✅ Period detected: {period_info['source_text']}")
            print(f"       📅 Years: {period_info['years_span']}")

            # [REST OF THE FUNCTION CONTINUES WITH EXISTING LOGIC BUT WITH ENHANCED FILTERING]
            # Find date headers
            date_headers = []
            date_row = None
            date_start_col = None

            for row_num in range(section['start_row'], min(section['start_row'] + 15, sheet.max_row + 1)):
                for col_num in range(3, min(50, sheet.max_column + 1)):
                    try:
                        cell_value = sheet.cell(row=row_num, column=col_num).value
                        if cell_value and DATE_PATTERN.match(str(cell_value)):
                            date_row = row_num
                            date_start_col = col_num

                            # Collect consecutive date headers
                            for date_col in range(col_num, min(sheet.max_column + 1, col_num + 500)):
                                date_cell = sheet.cell(row=row_num, column=date_col).value
                                if date_cell and DATE_PATTERN.match(str(date_cell)):
                                    date_headers.append(str(date_cell))
                                else:
                                    if len(date_headers) > 10:
                                        break
                            break
                    except:
                        continue
                if date_headers:
                    break

            if not date_headers:
                print(f"    ❌ No date headers found")
                continue

            print(f"    📅 Found {len(date_headers)} date headers")

            year_headers = assign_years_to_dates_improved(date_headers, period_info, debug=True)

            # APPLY FILTERING TO DATE HEADERS DURING EXTRACTION
            if active_filter:
                filtered_date_headers = []
                filtered_year_headers = []

                for i, (date_header, year_header) in enumerate(zip(date_headers, year_headers)):
                    include_date = False

                    if active_filter['type'] == 'year_range':
                        if year_header and year_header.isdigit():
                            year_int = int(year_header)
                            include_date = active_filter['start_year'] <= year_int <= active_filter['end_year']

                    elif active_filter['type'] == 'month_range':
                        if year_header and year_header.isdigit():
                            parsed_date = parse_date_for_month_filter(date_header, year_header)
                            if parsed_date:
                                include_date = date_falls_in_month_filter(parsed_date, active_filter)

                    elif active_filter['type'] == 'day_range':  # NEW DAY FILTER LOGIC
                        if year_header and year_header.isdigit():
                            parsed_date = parse_date_for_day_filter(date_header, year_header)
                            if parsed_date:
                                include_date = date_falls_in_day_filter(parsed_date, active_filter)

                    if include_date:
                        filtered_date_headers.append(date_header)
                        filtered_year_headers.append(year_header)

                # Update to use only filtered headers
                original_count = len(date_headers)
                date_headers = filtered_date_headers
                year_headers = filtered_year_headers

                print(f"    📅 Date headers filtered: {original_count} → {len(date_headers)} ({filter_description})")

                if not date_headers:
                    print(f"    ⏭️ Skipping section - no dates in filter range")
                    continue

            # Find data start row
            data_start_row = date_row + 1 if date_row else section['start_row'] + 3

            for search_row in range(data_start_row, min(section['start_row'] + 15, sheet.max_row + 50)):
                try:
                    cell_value_a = sheet.cell(row=search_row, column=1).value
                    cell_value_b = sheet.cell(row=search_row, column=2).value

                    if ((cell_value_a and isinstance(cell_value_a, str) and
                         any(keyword in cell_value_a for keyword in HOSPITAL_KEYWORDS)) or
                            (cell_value_b and isinstance(cell_value_b, str) and
                             any(keyword in cell_value_b for keyword in HOSPITAL_KEYWORDS))):
                        data_start_row = search_row
                        break
                except:
                    continue

            # Extract data rows (now only processing filtered date columns)
            current_hospital = None
            current_method_category = None
            end_row = min(data_start_row + 500, sheet.max_row + 1)
            records_extracted = 0

            for row_num in range(data_start_row, end_row):
                try:
                    # OPTIMIZATION: Only read columns we need (based on filtered date headers)
                    max_col = min(len(date_headers) + date_start_col + 10, sheet.max_column + 1)
                    row_values = [sheet.cell(row=row_num, column=col).value
                                  for col in range(1, max_col)]

                    if not any(row_values[:4]):
                        continue

                    if should_skip_row(row_values):
                        continue

                    hospital_name = find_hospital_name(row_values)
                    if hospital_name:
                        current_hospital = hospital_name

                    if not current_hospital:
                        continue

                    # Category extraction logic (same as before)
                    if section['type'] == 'method_patient_class':
                        if hospital_name and len(row_values) > 1 and row_values[1]:
                            current_method_category = str(row_values[1]).strip()

                        method_or_category = current_method_category
                        patient_class = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else None
                        category_type = "Method"

                    elif section['type'] == 'directorate':
                        method_or_category = str(row_values[2]).strip() if len(row_values) > 2 and row_values[
                            2] else None
                        patient_class = "Emergency Inpatient"
                        category_type = "Directorate"

                    elif section['type'] == 'diagnosis':
                        method_or_category = str(row_values[2]).strip() if len(row_values) > 2 and row_values[
                            2] else None
                        patient_class = "Emergency Inpatient"
                        category_type = "Primary Diagnosis"

                    else:
                        method_or_category = str(row_values[1]).strip() if len(row_values) > 1 and row_values[
                            1] else None
                        patient_class = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else None
                        category_type = "Unknown"

                    if (not method_or_category or not patient_class or not method_or_category.strip()):
                        continue

                    method_text = str(method_or_category).strip().lower()
                    patient_text = str(patient_class).strip().lower()

                    if any(indicator in method_text or indicator in patient_text for indicator in SKIP_KEYWORDS):
                        continue

                    # Create record
                    record = {
                        'hospital': current_hospital,
                        'section': section['title'],
                        'category_type': category_type,
                        'category_value': method_or_category,
                        'patient_class': patient_class,
                        'sheet_name': sheet_name,
                        'file_name': file_context['file_name'],
                        'period_info': period_info,
                        'daily_data': {}
                    }

                    # EFFICIENT: Only extract daily values for filtered dates
                    data_points_added = 0
                    for i, (date, year) in enumerate(zip(date_headers, year_headers)):
                        col_index = date_start_col - 1 + i
                        if col_index < len(row_values):
                            value = row_values[col_index]
                            if value is not None and str(value).strip():
                                try:
                                    if isinstance(value, str):
                                        value = value.replace(',', '').replace(' ', '')
                                    num_value = int(float(value))

                                    date_key = f"{date}-{year}" if year and year != 'Unknown' else date
                                    record['daily_data'][date_key] = num_value
                                    data_points_added += 1
                                except (ValueError, TypeError):
                                    pass

                    if record['daily_data']:
                        all_structured_data.append(record)
                        records_extracted += 1

                except Exception:
                    continue

            print(f"    ✅ Extracted {records_extracted} records from this section")

    workbook.close()
    total_data_points = sum(len(record['daily_data']) for record in all_structured_data)
    print(f"  ✅ Total extracted: {len(all_structured_data)} records with {total_data_points} data points")
    print(f"  📅 Filter applied: {filter_description}")

    return all_structured_data


def detect_cross_file_overlaps(all_files_data: Dict[str, List[Dict[str, Any]]], debug: bool = True) -> Dict[str, Any]:
    """Detect overlaps across multiple files"""
    if debug:
        print(f"\n🔍 DETECTING CROSS-FILE OVERLAPS:")

    overlap_analysis = {
        'total_data_points': 0,
        'unique_data_points': 0,
        'cross_file_overlaps': [],
        'date_collisions': [],
        'period_overlaps': [],
        'signature_to_files': defaultdict(list),
        'signature_to_values': defaultdict(list)
    }

    all_signatures = set()
    file_periods = defaultdict(dict)
    all_dates = set()

    # Collect signatures and periods
    for file_name, records in all_files_data.items():
        file_dates = set()

        for record in records:
            for date_key, value in record['daily_data'].items():
                signature = create_date_key_signature(
                    date_key, record['hospital'], record['category_type'],
                    record['category_value'], record['patient_class']
                )

                overlap_analysis['total_data_points'] += 1
                all_signatures.add(signature)

                overlap_analysis['signature_to_files'][signature].append(file_name)
                overlap_analysis['signature_to_values'][signature].append(value)

                file_dates.add(date_key)

        if file_dates:
            file_periods[file_name] = {
                'start_date': min(file_dates),
                'end_date': max(file_dates),
                'total_dates': len(file_dates),
                'date_set': file_dates
            }
            all_dates.update(file_dates)

    overlap_analysis['unique_data_points'] = len(all_signatures)

    # Detect overlaps
    for signature, files in overlap_analysis['signature_to_files'].items():
        if len(files) > 1:
            values = overlap_analysis['signature_to_values'][signature]
            conflict_severity = 'same_values' if len(set(values)) == 1 else 'conflicting_values'

            overlap_analysis['cross_file_overlaps'].append({
                'signature': signature,
                'files_involved': sorted(set(files)),
                'values': values,
                'conflict_severity': conflict_severity
            })

    # Detect date collisions
    for signature, values in overlap_analysis['signature_to_values'].items():
        if len(set(values)) > 1:
            overlap_analysis['date_collisions'].append({
                'signature': signature,
                'conflicting_values': values
            })

    # Detect period overlaps
    file_names = list(file_periods.keys())
    for i in range(len(file_names)):
        for j in range(i + 1, len(file_names)):
            file1, file2 = file_names[i], file_names[j]

            overlap_dates = file_periods[file1]['date_set'].intersection(file_periods[file2]['date_set'])
            if overlap_dates:
                file1_total = file_periods[file1]['total_dates']
                file2_total = file_periods[file2]['total_dates']

                overlap_analysis['period_overlaps'].append({
                    'file1': file1,
                    'file2': file2,
                    'overlapping_dates': sorted(list(overlap_dates)),
                    'overlap_count': len(overlap_dates),
                    'file1_period': f"{file_periods[file1]['start_date']} to {file_periods[file1]['end_date']}",
                    'file2_period': f"{file_periods[file2]['start_date']} to {file_periods[file2]['end_date']}",
                    'overlap_percentage_file1': (len(overlap_dates) / file1_total * 100) if file1_total > 0 else 0,
                    'overlap_percentage_file2': (len(overlap_dates) / file2_total * 100) if file2_total > 0 else 0
                })

    if debug:
        print(f"  📊 Analysis Results:")
        print(f"    🔢 Total data points: {overlap_analysis['total_data_points']}")
        print(f"    🔑 Unique signatures: {overlap_analysis['unique_data_points']}")
        print(f"    🔄 Cross-file overlaps: {len(overlap_analysis['cross_file_overlaps'])}")
        print(f"    📅 Date collisions: {len(overlap_analysis['date_collisions'])}")
        print(f"    📊 Period overlaps: {len(overlap_analysis['period_overlaps'])}")

    return overlap_analysis


def resolve_cross_file_overlaps(all_files_data: Dict[str, List[Dict[str, Any]]],
                                overlap_analysis: Dict[str, Any],
                                resolution_strategy: str = 'prefer_latest_file',
                                debug: bool = True) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Resolve cross-file overlaps based on specified strategy"""
    if debug:
        print(f"\n🔧 RESOLVING CROSS-FILE OVERLAPS:")
        print(f"  🎯 Strategy: {resolution_strategy}")
        print(f"  🔄 Overlaps to resolve: {len(overlap_analysis['cross_file_overlaps'])}")

    resolved_data = []
    resolution_report = {
        'strategy_used': resolution_strategy,
        'overlaps_resolved': 0,
        'conflicts_found': 0,
        'data_points_removed': 0,
        'resolution_details': []
    }

    processed_signatures = set()

    # Process each file's data
    for file_name, records in all_files_data.items():
        file_records_added = 0
        file_data_points_skipped = 0

        for record in records:
            new_daily_data = {}

            for date_key, value in record['daily_data'].items():
                signature = create_date_key_signature(
                    date_key, record['hospital'], record['category_type'],
                    record['category_value'], record['patient_class']
                )

                # Check if this signature has overlaps
                overlap_found = False
                for overlap in overlap_analysis['cross_file_overlaps']:
                    if signature.startswith(overlap['signature']):
                        overlap_found = True

                        if signature not in processed_signatures:
                            # First occurrence of this overlap - resolve it
                            should_keep = False

                            if resolution_strategy == 'prefer_latest_file':
                                files_involved = overlap['files_involved']
                                preferred_file = max(files_involved)
                                should_keep = (file_name == preferred_file)

                            elif resolution_strategy == 'prefer_first_file':
                                files_involved = overlap['files_involved']
                                preferred_file = min(files_involved)
                                should_keep = (file_name == preferred_file)

                            elif resolution_strategy == 'keep_highest_value':
                                should_keep = (value == max(overlap['values']))

                            elif resolution_strategy == 'average_values':
                                if overlap['conflict_severity'] == 'conflicting_values':
                                    avg_value = sum(overlap['values']) / len(overlap['values'])
                                    new_daily_data[date_key] = round(avg_value)
                                    processed_signatures.add(signature)
                                    resolution_report['conflicts_found'] += 1
                                    should_keep = False
                                else:
                                    should_keep = True

                            if should_keep:
                                new_daily_data[date_key] = value
                                processed_signatures.add(signature)

                                resolution_report['resolution_details'].append({
                                    'signature': signature[:12],
                                    'date_key': date_key,
                                    'category_value': record['category_value'],
                                    'files_involved': overlap['files_involved'],
                                    'chosen_file': file_name,
                                    'value_kept': value,
                                    'reason': f'{resolution_strategy.replace("_", " ").title()} strategy - chose {file_name}'
                                })
                            else:
                                file_data_points_skipped += 1
                                resolution_report['data_points_removed'] += 1

                            resolution_report['overlaps_resolved'] += 1
                        break

                if not overlap_found:
                    new_daily_data[date_key] = value

            # Create resolved record
            if new_daily_data:
                resolved_record = record.copy()
                resolved_record['daily_data'] = new_daily_data
                resolved_record['cross_file_resolution'] = {
                    'original_data_points': len(record['daily_data']),
                    'resolved_data_points': len(new_daily_data),
                    'points_removed': len(record['daily_data']) - len(new_daily_data)
                }
                resolved_data.append(resolved_record)
                file_records_added += 1

        if debug:
            print(f"    📁 {file_name}: {file_records_added} records, {file_data_points_skipped} data points removed")

    if debug:
        print(f"  ✅ Resolution complete:")
        print(f"    🔄 Overlaps resolved: {resolution_report['overlaps_resolved']}")
        print(f"    ⚠️ Conflicts found: {resolution_report['conflicts_found']}")
        print(f"    🗑️ Data points removed: {resolution_report['data_points_removed']}")

    return resolved_data, resolution_report


def create_flattened_record(record: Dict[str, Any], date_year: str, value: int) -> Dict[str, Any]:
    """Create a flattened record for Excel output with renamed columns"""
    base_record = {
        'Hospital_Name': record['hospital'],
        'Section': record['section'],
        'Category_Name': record['category_type'],
        'Category_Value': record['category_value'],
        'Patient_Class': record['patient_class'],
        'Sheet_Name': record['sheet_name'],
        'File_Name': record['file_name'],
        'Period_Source': record.get('period_info', {}).get('source_text', 'Unknown')
    }

    # Initialize default values
    day_num = 0
    month_abbr = 'Unknown'
    year_part = 'Unknown'

    # Parse date and year - Handle format like "1-Sep-2021" or "1-Sep"
    if '-202' in date_year:  # Contains year like "1-Sep-2021"
        parts = date_year.rsplit('-', 1)  # Split from right to separate year
        if len(parts) == 2 and len(parts[1]) == 4 and parts[1].isdigit():
            date_part = parts[0]  # "1-Sep"
            year_part = parts[1]  # "2021"
            base_record['Year'] = year_part

            # Convert date format from "1-Sep" to "01/09/2021"
            if '-' in date_part:
                day_month = date_part.split('-')
                if len(day_month) == 2:
                    day_num = int(day_month[0]) if day_month[0].isdigit() else 1
                    day = day_month[0].zfill(2)  # "01"
                    month_abbr = day_month[1]  # "Sep"
                    month_num = MONTH_NUMBERS.get(month_abbr, '00')
                    base_record['Date'] = f"{day}/{month_num}/{year_part}"

                    # Set month info
                    base_record['Month_num'] = month_num
                    base_record['Month'] = MONTH_MAPPING.get(month_abbr, 'Unknown')
                else:
                    base_record['Date'] = date_year
                    base_record['Month_num'] = '00'
                    base_record['Month'] = 'Unknown'
            else:
                base_record['Date'] = date_year
                base_record['Month_num'] = '00'
                base_record['Month'] = 'Unknown'
        else:
            base_record['Year'] = 'Unknown'
            base_record['Date'] = date_year
            base_record['Month_num'] = '00'
            base_record['Month'] = 'Unknown'
    else:  # No year like "1-Sep"
        base_record['Year'] = 'Unknown'

        # Handle dates without year - try to format them
        if '-' in date_year:
            day_month = date_year.split('-')
            if len(day_month) == 2:
                day_num = int(day_month[0]) if day_month[0].isdigit() else 1
                day = day_month[0].zfill(2)  # "01"
                month_abbr = day_month[1]  # "Sep"
                month_num = MONTH_NUMBERS.get(month_abbr, '00')
                base_record['Date'] = f"{day}/{month_num}/YYYY"

                # Set month info
                base_record['Month_num'] = month_num
                base_record['Month'] = MONTH_MAPPING.get(month_abbr, 'Unknown')
            else:
                base_record['Date'] = date_year
                base_record['Month_num'] = '00'
                base_record['Month'] = 'Unknown'
        else:
            base_record['Date'] = date_year
            base_record['Month_num'] = '00'
            base_record['Month'] = 'Unknown'

    # Add new time-based features with renamed columns
    base_record['Season'] = SEASON_MAPPING.get(month_abbr, 'Unknown')
    base_record['Quarter'] = QUARTER_MAPPING.get(month_abbr, 'Unknown')

    # Get day of week information with renamed columns
    day_info = get_day_of_week_info(day_num, month_abbr, year_part)
    base_record['Day_of_Week'] = day_info['day_of_week']
    base_record['Day_of_week_num'] = day_info['day_of_week_num']
    base_record['Is_Weekend'] = day_info['is_weekend']

    base_record['No_of_Admissions'] = value
    return base_record


def apply_year_filter_to_dataframe(df: pd.DataFrame, year_filter: tuple) -> pd.DataFrame:
    """
    Apply year filter to a processed DataFrame

    Args:
        df: DataFrame to filter
        year_filter: Tuple of (start_year, end_year)

    Returns:
        Filtered DataFrame
    """
    if year_filter is None or df.empty:
        return df

    original_count = len(df)

    # Filter by Year column if it exists
    if 'Year' in df.columns:
        # Handle both string and numeric years
        df['Year_Numeric'] = pd.to_numeric(df['Year'], errors='coerce')
        df = df[
            (df['Year_Numeric'] >= year_filter[0]) &
            (df['Year_Numeric'] <= year_filter[1]) &
            (df['Year_Numeric'].notna())
            ]
        df = df.drop('Year_Numeric', axis=1)

    # Also filter by Date column if Year column filtering didn't work well
    elif 'Date' in df.columns:
        try:
            # Parse dates and extract years
            df['Date_Parsed'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
            df['Date_Year'] = df['Date_Parsed'].dt.year
            df = df[
                (df['Date_Year'] >= year_filter[0]) &
                (df['Date_Year'] <= year_filter[1]) &
                (df['Date_Year'].notna())
                ]
            df = df.drop(['Date_Parsed', 'Date_Year'], axis=1)
        except Exception as e:
            print(f"⚠️ Could not filter by date: {e}")

    filtered_count = len(df)
    if original_count != filtered_count:
        print(f"📅 DataFrame year filter applied: {original_count:,} → {filtered_count:,} records")

    return df


def save_resolved_data_to_excel(resolved_data: List[Dict[str, Any]],
                                resolution_report: Dict[str, Any],
                                output_path: str,
                                bank_holidays_df: pd.DataFrame = None,
                                enable_weather: bool = False,
                                year_filter: tuple = None,
                                month_filter: Dict[str, Any] = None,
                                day_filter: Dict[str, Any] = None) -> str:
    """Save the cross-file resolved data to Excel with PROPER SORTING based on filter criteria"""
    print(f"\n💾 SAVING CROSS-FILE RESOLVED DATA WITH ENHANCED SORTING:")

    # Enhanced filter suffix logic
    filter_suffix = ""
    active_filter_desc = "No filter"

    if day_filter:
        start_str = day_filter['start_date'].strftime('%Y%m%d')
        end_str = day_filter['end_date'].strftime('%Y%m%d')
        filter_suffix = f"_day_{start_str}_to_{end_str}"
        active_filter_desc = f"Day filter: {day_filter['description']}"
    elif month_filter:
        safe_desc = month_filter['description'].replace(' ', '_').replace('-', '_')
        filter_suffix = f"_{safe_desc}"
        active_filter_desc = f"Month filter: {month_filter['description']}"
    elif year_filter:
        filter_suffix = f"_{year_filter[0]}_{year_filter[1]}"
        active_filter_desc = f"Year filter: {year_filter[0]}-{year_filter[1]}"

    # Create output filename
    if enable_weather:
        output_filename = f"Admission_Data_with_Weather{filter_suffix}.xlsx"
    else:
        output_filename = f"Admission_Data{filter_suffix}.xlsx"

    excel_path = os.path.join(output_path, output_filename)

    # Column definitions
    base_columns = ['Hospital_Name', 'Section', 'Category_Name', 'Category_Value', 'Patient_Class',
                    'Sheet_Name', 'File_Name', 'Period_Source',
                    'Year', 'Month_num', 'Month', 'Season', 'Quarter',
                    'Date', 'Day_of_Week', 'Day_of_week_num', 'Is_Weekend', 'No_of_Admissions']

    full_column_order = base_columns[:-1] + ['Is_Bank_Holiday', 'Bank_Holiday_Name', 'No_of_Admissions']

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        sheets_created = 0

        # SHEET 1: Method/Patient Class data with enhanced sorting
        method_patient_class_data = [record for record in resolved_data
                                     if "Method/Patient Class" in record['section']]

        if method_patient_class_data:
            try:
                flattened_data = []
                for record in method_patient_class_data:
                    for date_year, value in record['daily_data'].items():
                        flattened_record = create_flattened_record(record, date_year, value)
                        flattened_data.append(flattened_record)

                if flattened_data:
                    df = pd.DataFrame(flattened_data)

                    # Ensure all base columns exist
                    for col in base_columns:
                        if col not in df.columns:
                            if col == 'No_of_Admissions':
                                df[col] = 0
                            else:
                                df[col] = 'Unknown'

                    # Select base columns first
                    df = df[base_columns]

                    # Add bank holiday information if available
                    if bank_holidays_df is not None and not bank_holidays_df.empty:
                        df = add_bank_holiday_info(df, bank_holidays_df)
                        df = df[full_column_order]
                    else:
                        df['Is_Bank_Holiday'] = 'No'
                        df['Bank_Holiday_Name'] = None
                        df = df[full_column_order]

                    if enable_weather:
                        df = enhance_dataframe_with_weather_data(df)

                    # ⭐ ENHANCED SORTING - This is the key improvement! ⭐
                    print(f"  🔄 Applying enhanced sorting for Method/Patient Class data...")
                    df = sort_dataframe_by_filter_criteria(df, year_filter, month_filter, day_filter)

                    df.to_excel(writer, sheet_name='Method_Patient_Class', index=False)
                    sheets_created += 1

                    # Show sorting confirmation with sample data
                    print(f"  📊 Method/Patient Class sheet: {len(df)} records")
                    print(f"     📅 Sorted according to: {active_filter_desc}")

                    # Show first few dates to confirm sorting
                    if len(df) > 0 and 'Date' in df.columns:
                        sample_dates = df['Date'].head(5).tolist()
                        print(f"     📋 First 5 dates after sorting: {sample_dates}")

            except Exception as e:
                print(f"  ⚠️ Error creating Method/Patient Class sheet: {e}")

        # SHEET 2: Emergency IP tables with enhanced sorting
        emergency_data = [record for record in resolved_data
                          if "Emergency IP Admissions by Admitting Directorate" in record['section'] or
                          "Emergency IP Admissions by Admitting Primary Diagnosis" in record['section']]

        if emergency_data:
            try:
                flattened_data = []
                for record in emergency_data:
                    for date_year, value in record['daily_data'].items():
                        flattened_record = create_flattened_record(record, date_year, value)
                        flattened_data.append(flattened_record)

                if flattened_data:
                    df = pd.DataFrame(flattened_data)

                    # Ensure all base columns exist
                    for col in base_columns:
                        if col not in df.columns:
                            if col == 'No_of_Admissions':
                                df[col] = 0
                            else:
                                df[col] = 'Unknown'

                    # Select base columns first
                    df = df[base_columns]

                    # Add bank holiday information if available
                    if bank_holidays_df is not None and not bank_holidays_df.empty:
                        df = add_bank_holiday_info(df, bank_holidays_df)
                        df = df[full_column_order]
                    else:
                        df['Is_Bank_Holiday'] = 'No'
                        df['Bank_Holiday_Name'] = None
                        df = df[full_column_order]

                    if enable_weather:
                        df = enhance_dataframe_with_weather_data(df)

                    # ⭐ ENHANCED SORTING - This is the key improvement! ⭐
                    print(f"  🔄 Applying enhanced sorting for Emergency IP data...")
                    df = sort_dataframe_by_filter_criteria(df, year_filter, month_filter, day_filter)

                    df.to_excel(writer, sheet_name='Emergency_IP', index=False)
                    sheets_created += 1

                    # Show sorting confirmation with sample data
                    print(f"  📊 Emergency IP sheet: {len(df)} records")
                    print(f"     📅 Sorted according to: {active_filter_desc}")

                    # Show first few dates to confirm sorting
                    if len(df) > 0 and 'Date' in df.columns:
                        sample_dates = df['Date'].head(5).tolist()
                        print(f"     📋 First 5 dates after sorting: {sample_dates}")

            except Exception as e:
                print(f"  ⚠️ Error creating Emergency IP sheet: {e}")

        # If no sheets were created, create a status sheet
        if sheets_created == 0:
            print(f"  ⚠️ No data sheets created - creating minimal status sheet")
            status_data = {
                'Status': ['No data sheets created', 'Check filter criteria'],
                'Records_Found': [len(resolved_data), 'N/A'],
                'Filter_Applied': [active_filter_desc, 'N/A']
            }
            status_df = pd.DataFrame(status_data)
            status_df.to_excel(writer, sheet_name='Status', index=False)

    print(f"  ✅ Admission data Excel saved: {output_filename}")
    print(f"  📄 Sheets created: {sheets_created}")
    print(f"  🔄 Data sorted according to: {active_filter_desc}")

    if bank_holidays_df is not None and not bank_holidays_df.empty:
        print(f"  📅 Enhanced with bank holiday information")
    if enable_weather:
        print(f"  🌤️ Enhanced with weather information")

    return excel_path


def process_admission_files_with_weather(file_paths: List[str], output_directory: str, enable_weather: bool) -> str:
    return main_efficient(file_paths, output_directory, enable_weather=enable_weather)


def main_efficient(excel_file_paths: List[str],
                   output_path: str,
                   update_progress=None,
                   enable_weather: bool = False,
                   year_filter: tuple = None):
    """
    EFFICIENT VERSION: Main execution function with year filtering during extraction
    """
    print("=" * 80)
    print("🎯 EFFICIENT ADMISSIONS DATA EXTRACTOR")
    if year_filter:
        print(f"📅 YEAR FILTER: {year_filter[0]} - {year_filter[1]} (applied during extraction)")
    print("=" * 80)

    RESOLUTION_STRATEGY = 'prefer_latest_file'
    FETCH_BANK_HOLIDAYS = True
    UK_REGION = "england-and-wales"

    existing_files = [path for path in excel_file_paths if os.path.exists(path)]

    if not existing_files:
        print("❌ ERROR: No Excel files found!")
        if update_progress:
            update_progress(0, "Error: No valid Excel files found")
        raise ValueError("No valid Excel files found")

    print(f"🎯 Processing {len(existing_files)} files:")
    for i, file_path in enumerate(existing_files, 1):
        print(f"  {i}. {os.path.basename(file_path)}")

    print(f"\n🔧 Configuration:")
    print(f"  📊 Resolution strategy: {RESOLUTION_STRATEGY}")
    print(f"  📂 Output location: {output_path}")
    print(f"  🏛️ Bank holidays: {'Enabled' if FETCH_BANK_HOLIDAYS else 'Disabled'}")
    print(f"  🌤️ Weather integration: {'Enabled' if enable_weather else 'Disabled'}")
    if year_filter:
        print(f"  📅 Year filter: {year_filter[0]} to {year_filter[1]} ✅ (EFFICIENT MODE)")
    else:
        print(f"  📅 Year filter: Disabled (all years)")

    try:
        # Step 1: Extract data with efficient year filtering
        if update_progress:
            update_progress(1, "Extracting data from Excel files (with efficient year filtering)...")

        print(f"\n" + "=" * 80)
        print("📊 STEP 1: EFFICIENT DATA EXTRACTION WITH YEAR FILTERING")
        print("=" * 80)

        all_files_data = {}
        extraction_summary = {}
        all_years = set()

        for file_path in existing_files:
            print(f"\n📁 Processing: {os.path.basename(file_path)}")

            # FIX: Use the correct function name
            structured_data = extract_basic_admissions_data_with_enhanced_filter(file_path, year_filter, None)

            if structured_data:
                all_files_data[os.path.basename(file_path)] = structured_data
                extraction_summary[os.path.basename(file_path)] = {
                    'records': len(structured_data),
                    'total_data_points': sum(len(record['daily_data']) for record in structured_data)
                }

                # Collect years for bank holiday fetching
                for record in structured_data:
                    period_info = record.get('period_info', {})
                    years_span = period_info.get('years_span', [])
                    for year in years_span:
                        if year.isdigit():
                            year_int = int(year)
                            if not year_filter or (year_filter[0] <= year_int <= year_filter[1]):
                                all_years.add(year_int)

                print(
                    f"  ✅ SUCCESS: {len(structured_data):,} records, {extraction_summary[os.path.basename(file_path)]['total_data_points']:,} data points")
            else:
                msg = "No data in specified year range" if year_filter else "No data extracted"
                print(f"  ⚠️ WARNING: {msg} from {os.path.basename(file_path)}")

        if not all_files_data:
            error_msg = "No data extracted from any files in the specified year range!" if year_filter else "No data extracted from any files!"
            print(f"❌ ERROR: {error_msg}")
            if update_progress:
                update_progress(1, f"Error: {error_msg}")
            raise ValueError(error_msg)

        # Step 1.5: Fetch bank holidays
        if update_progress:
            update_progress(2, "Fetching bank holidays...")

        bank_holidays_df = pd.DataFrame()
        if FETCH_BANK_HOLIDAYS and all_years:
            print(f"\n" + "=" * 80)
            print("🏛️ STEP 1.5: FETCHING BANK HOLIDAYS")
            print("=" * 80)

            years_list = sorted(list(all_years))
            print(f"🗓️ Years detected in filtered data: {years_list}")

            try:
                bank_holidays_df = fetch_bank_holidays(years_list, UK_REGION)
                if not bank_holidays_df.empty:
                    print(f"✅ Successfully fetched {len(bank_holidays_df)} bank holidays")
                else:
                    print("⚠️ No bank holidays fetched")
            except Exception as e:
                print(f"⚠️ Failed to fetch bank holidays: {e}")

        # Step 2: Detect cross-file overlaps
        if update_progress:
            update_progress(3, "Detecting cross-file overlaps...")

        print(f"\n" + "=" * 80)
        print("🔍 STEP 2: DETECTING CROSS-FILE OVERLAPS")
        print("=" * 80)

        overlap_analysis = detect_cross_file_overlaps(all_files_data, debug=True)

        # Step 3: Resolve overlaps
        if update_progress:
            update_progress(4, "Resolving cross-file overlaps...")

        print(f"\n" + "=" * 80)
        print("🔧 STEP 3: RESOLVING CROSS-FILE OVERLAPS")
        print("=" * 80)

        resolved_data, resolution_report = resolve_cross_file_overlaps(
            all_files_data, overlap_analysis, RESOLUTION_STRATEGY, debug=True)

        # Step 4: Save Excel output (no additional year filtering needed!)
        if update_progress:
            update_progress(5, "Saving Excel output...")

        print(f"\n" + "=" * 80)
        print("💾 STEP 4: SAVING EXCEL OUTPUT")
        print("=" * 80)

        # Save resolved Excel data - NO ADDITIONAL YEAR FILTERING NEEDED
        excel_path = save_resolved_data_to_excel(
            resolved_data,
            resolution_report,
            output_path,
            bank_holidays_df,
            enable_weather=enable_weather,
            year_filter=None  # Set to None since filtering already done
        )

        if excel_path and os.path.exists(excel_path):
            print(f"✅ File successfully created: {excel_path}")
        else:
            error_msg = f"❌ File creation failed: {excel_path}"
            print(error_msg)
            if update_progress:
                update_progress(0, error_msg)
            raise ValueError(error_msg)

        # Final summary
        print(f"\n" + "=" * 80)
        print("🎉 EFFICIENT PROCESSING COMPLETE")
        print("=" * 80)

        if year_filter:
            print(f"📅 EFFICIENT YEAR FILTERING APPLIED: {year_filter[0]} - {year_filter[1]}")
            print(f"   Filtering was done DURING extraction, not after!")

        print(f"\n📊 EXTRACTION RESULTS:")
        for file_name, summary in extraction_summary.items():
            print(f"  📁 {file_name}:")
            print(f"    📋 Records: {summary['records']:,}")
            print(f"    📊 Data points: {summary['total_data_points']:,}")

        print(f"\n🎉 EFFICIENT ADMISSIONS DATA EXTRACTION COMPLETE!")
        if year_filter:
            print(f"⚡ Year filter {year_filter[0]}-{year_filter[1]} applied efficiently during extraction!")

        if update_progress:
            update_progress(5, "Processing completed successfully!")

        return excel_path

    except Exception as e:
        error_msg = f"ERROR during processing: {str(e)}"
        print(f"\n❌ {error_msg}")
        if update_progress:
            update_progress(0, f"Error: {str(e)}")
        raise e


# Backward compatibility - keep the original main function name but use efficient version
def main(excel_file_paths: List[str],
         output_path: str,
         update_progress=None,
         enable_weather: bool = False,
         year_filter: tuple = None):
    """
    Main function - now uses the efficient version by default
    """
    return main_efficient(excel_file_paths, output_path, update_progress, enable_weather, year_filter)


def main_with_month_filter(excel_file_paths: List[str],
                           output_path: str,
                           update_progress=None,
                           enable_weather: bool = False,
                           month_filter: Dict[str, Any] = None):
    """
    Main function specifically for month filtering
    """
    print("=" * 80)
    print("🎯 ADMISSIONS DATA EXTRACTOR WITH MONTH FILTERING")
    print(f"📅 MONTH FILTER: {month_filter['description'] if month_filter else 'None'}")
    print("=" * 80)

    # Use the enhanced extraction function
    RESOLUTION_STRATEGY = 'prefer_latest_file'
    FETCH_BANK_HOLIDAYS = True
    UK_REGION = "england-and-wales"

    existing_files = [path for path in excel_file_paths if os.path.exists(path)]

    if not existing_files:
        print("❌ ERROR: No Excel files found!")
        if update_progress:
            update_progress(0, "Error: No valid Excel files found")
        raise ValueError("No valid Excel files found")

    # Extract data with month filtering
    all_files_data = {}
    all_years = set()

    for file_path in existing_files:
        print(f"\n📁 Processing: {os.path.basename(file_path)}")

        # Use the enhanced extraction function
        structured_data = extract_basic_admissions_data_with_enhanced_filter(
            file_path, year_filter=None, month_filter=month_filter
        )

        if structured_data:
            all_files_data[os.path.basename(file_path)] = structured_data

            # Collect years for bank holiday fetching
            for record in structured_data:
                period_info = record.get('period_info', {})
                years_span = period_info.get('years_span', [])
                for year in years_span:
                    if year.isdigit():
                        year_int = int(year)
                        if month_filter:
                            filter_years = range(month_filter['start_year'], month_filter['end_year'] + 1)
                            if year_int in filter_years:
                                all_years.add(year_int)

    # Continue with bank holidays, overlap detection, and saving
    bank_holidays_df = pd.DataFrame()
    if FETCH_BANK_HOLIDAYS and all_years:
        years_list = sorted(list(all_years))
        try:
            bank_holidays_df = fetch_bank_holidays(years_list, UK_REGION)
        except Exception as e:
            print(f"⚠️ Failed to fetch bank holidays: {e}")

    # Detect and resolve overlaps
    overlap_analysis = detect_cross_file_overlaps(all_files_data, debug=True)
    resolved_data, resolution_report = resolve_cross_file_overlaps(
        all_files_data, overlap_analysis, RESOLUTION_STRATEGY, debug=True)

    # Save Excel output with month filter info
    excel_path = save_resolved_data_to_excel(
        resolved_data,
        resolution_report,
        output_path,
        bank_holidays_df,
        enable_weather=enable_weather,
        year_filter=None,
        month_filter=month_filter
    )

    return excel_path

def main_with_day_filter(excel_file_paths: List[str],
                         output_path: str,
                         update_progress=None,
                         enable_weather: bool = False,
                         day_filter: Dict[str, Any] = None):
    """
    Main function specifically for day filtering
    """
    print("=" * 80)
    print("🎯 ADMISSIONS DATA EXTRACTOR WITH DAY FILTERING")
    if day_filter:
        print(f"📅 DAY FILTER: {day_filter['description']}")
        print(f"   Duration: {day_filter['duration_days']} days")
    else:
        print("📅 NO DAY FILTER")
    print("=" * 80)

    # Use the enhanced extraction function
    RESOLUTION_STRATEGY = 'prefer_latest_file'
    FETCH_BANK_HOLIDAYS = True
    UK_REGION = "england-and-wales"

    existing_files = [path for path in excel_file_paths if os.path.exists(path)]

    if not existing_files:
        print("❌ ERROR: No Excel files found!")
        if update_progress:
            update_progress(0, "Error: No valid Excel files found")
        raise ValueError("No valid Excel files found")

    # Extract data with day filtering
    all_files_data = {}
    all_years = set()

    for file_path in existing_files:
        print(f"\n📁 Processing: {os.path.basename(file_path)}")

        # Use the enhanced extraction function with day filter
        structured_data = extract_basic_admissions_data_with_enhanced_filter(
            file_path, year_filter=None, month_filter=None, day_filter=day_filter
        )

        if structured_data:
            all_files_data[os.path.basename(file_path)] = structured_data

            # Collect years for bank holiday fetching
            for record in structured_data:
                period_info = record.get('period_info', {})
                years_span = period_info.get('years_span', [])
                for year in years_span:
                    if year.isdigit():
                        year_int = int(year)
                        if day_filter:
                            if day_filter['start_year'] <= year_int <= day_filter['end_year']:
                                all_years.add(year_int)

    # Continue with bank holidays, overlap detection, and saving
    bank_holidays_df = pd.DataFrame()
    if FETCH_BANK_HOLIDAYS and all_years:
        years_list = sorted(list(all_years))
        try:
            bank_holidays_df = fetch_bank_holidays(years_list, UK_REGION)
        except Exception as e:
            print(f"⚠️ Failed to fetch bank holidays: {e}")

    # Detect and resolve overlaps
    overlap_analysis = detect_cross_file_overlaps(all_files_data, debug=True)
    resolved_data, resolution_report = resolve_cross_file_overlaps(
        all_files_data, overlap_analysis, RESOLUTION_STRATEGY, debug=True)

    # Save Excel output with day filter info
    excel_path = save_resolved_data_to_excel(
        resolved_data,
        resolution_report,
        output_path,
        bank_holidays_df,
        enable_weather=enable_weather,
        year_filter=None,
        month_filter=None,
        day_filter=day_filter  # ADD THIS PARAMETER
    )

    return excel_path


# FIX: Clean up the module ending
if __name__ == "__main__":
    print("🎯 Enhanced Admissions Data Processor with Weather Integration")
    print("=" * 60)
    print("✅ Core module loaded successfully!")
    print("✅ Features available:")
    print("   - Admissions data extraction and processing")
    print("   - Cross-file overlap detection and resolution")
    print("   - Dynamic bank holiday integration (GOV.UK API)")
    print("   - Weather data integration (Open-Meteo API)")
    print("   - Automatic location detection and geocoding")
    print("   - Enhanced data analysis capabilities")
    print("   - EFFICIENT year filtering during extraction")
    print("   - Month filtering support")
    print("\n⚡ EFFICIENCY IMPROVEMENTS:")
    print("   - Year filtering applied during extraction (not after)")
    print("   - Early section skipping for non-matching years")
    print("   - Optimized column reading based on filtered dates")
    print("   - Significantly faster processing for year-filtered queries")